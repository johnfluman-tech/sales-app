import pyodbc
import pandas as pd
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import time

# ── Google Sheets imports ─────────────────────────────────────────────────────
try:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    SHEETS_AVAILABLE = True
except ImportError:
    SHEETS_AVAILABLE = False
    print("WARNING: Google libraries not installed.")
    print("         Run: pip install google-auth google-api-python-client")
    print("         Continuing without Google Sheets sync.\n")

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
CREDENTIALS_FILE = r"C:\scripts\google_credentials.json"
SPREADSHEET_ID   = "1xH_OC_fvSwQWZet95xBlJ951LsaQWru-glv2rkSnDm4"
HISTORY_SHEET_ID = "192ELLBgiEUZ3z6ytkWXbmA24wS2BjHmclJdf_UF_c24"   # Intransit History Data
ACTIVITY_SHEET_ID = "12dJ0eLFQse-_pi1k6rXc25yTkwwYTjMm32v11tGKeRo"  # Intransit Activity Data
SCOPES           = ["https://www.googleapis.com/auth/spreadsheets"]

ACTIVE_REPS  = ['CKaren', 'BillP', 'PIan', 'RMauricio', 'LMancera', 'bcastor', 'FJohn', 'Anolan']
OUTPUT_DIR   = r"C:\scripts\reports"
TODAY        = date.today()
MONTH_STR    = TODAY.strftime("%Y-%m")
CUTOFF_DATE  = date(TODAY.year - 1, TODAY.month, TODAY.day)  # 1 year ago = inactive
NOTE_COLS     = ['REP_NOTE', 'FOLLOW_UP_DATE', 'REP_STATUS', 'REMOVE_FLAG', 'JOHN_APPROVAL']
PRESERVE_COLS = NOTE_COLS + ['HIDE_FROM_POOL']  # all columns to read back & preserve on overwrite

# Status thresholds
INACTIVE_DAYS       = 183   # 6 months — inactive
HIGH_VALUE_THRESH   = 100000
AT_RISK_THRESH      = 20000

# Colors
C_DARK_BLUE   = "1F3864"
C_MED_BLUE    = "2F5496"
C_LIGHT_BLUE  = "D9E1F2"
C_RED         = "C00000"
C_ORANGE      = "E26B0A"
C_BLACK       = "4A4A4A"
C_YELLOW      = "FFF2CC"
C_WHITE       = "FFFFFF"
C_LIGHT_GRAY  = "F2F2F2"

# ══════════════════════════════════════════════════════════════════════════════
#  GOOGLE SHEETS HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_sheets_service():
    if not SHEETS_AVAILABLE:
        return None
    if not os.path.exists(CREDENTIALS_FILE):
        print(f"  WARNING: Credentials not found at {CREDENTIALS_FILE} — skipping Sheets sync.")
        return None
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
        return build("sheets", "v4", credentials=creds)
    except Exception as e:
        print(f"  WARNING: Google auth failed — {e}")
        return None


def read_notes_from_sheet(service):
    """Read rep notes from every rep tab in the Sheet. Returns dict keyed by (customer_name, rep)."""
    notes = {}
    if service is None:
        return notes
    try:
        meta      = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        tab_names = [s['properties']['title'] for s in meta['sheets']]
        for tab in tab_names:
            if tab.startswith('_') or tab == 'MASTER':
                continue
            result = service.spreadsheets().values().get(
                spreadsheetId=SPREADSHEET_ID,
                range=f"'{tab}'!A1:Z5000"
            ).execute()
            rows = result.get('values', [])
            if len(rows) < 2:
                continue
            # Find the real header row (the one that contains 'CUSTOMER_NAME')
            # Rep tabs have 3 instruction/blank rows before the actual column headers
            header_idx = None
            for _i, _r in enumerate(rows):
                if 'CUSTOMER_NAME' in _r:
                    header_idx = _i
                    break
            if header_idx is None:
                continue
            headers = rows[header_idx]
            for row in rows[header_idx + 1:]:
                row += [''] * (len(headers) - len(row))
                rec = dict(zip(headers, row))
                key = (rec.get('CUSTOMER_NAME', '').strip(), rec.get('SALES_REP', '').strip())
                if not key[0]:
                    continue
                notes[key] = {c: rec.get(c, '') for c in PRESERVE_COLS}
        print(f"  → Read notes for {len(notes)} accounts from Google Sheet.")
    except Exception as e:
        print(f"  WARNING: Could not read Sheet notes — {e}")
    return notes


def ensure_tab_exists(service, tab_name, all_tabs, row_count=5000, sheet_id=None):
    """Create a Sheet tab if it doesn't already exist."""
    sid = sheet_id or SPREADSHEET_ID
    if tab_name not in all_tabs:
        rows = max(row_count + 100, 1000)
        cols = 20
        body = {"requests": [{"addSheet": {"properties": {
            "title": tab_name,
            "gridProperties": {"rowCount": rows, "columnCount": cols}
        }}}]}
        service.spreadsheets().batchUpdate(spreadsheetId=sid, body=body).execute()
        all_tabs.append(tab_name)


def get_all_tabs(service, sheet_id):
    """Return list of tab names for any spreadsheet."""
    try:
        meta = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        return [s['properties']['title'] for s in meta.get('sheets', [])]
    except Exception:
        return []


def backup_notes_to_sheet(service, all_tabs):
    """Before each run, snapshot all existing rep notes to _NOTES_BACKUP tab.
    Rows are only APPENDED — never deleted — so full history is always preserved."""
    if service is None:
        return
    try:
        ensure_tab_exists(service, '_NOTES_BACKUP', all_tabs)

        # Check if header exists
        existing = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range="'_NOTES_BACKUP'!A1:A1"
        ).execute()
        if not existing.get('values'):
            header = [['BACKUP_TIMESTAMP', 'CUSTOMER_NAME', 'SALES_REP'] + NOTE_COLS]
            service.spreadsheets().values().update(
                spreadsheetId=SPREADSHEET_ID,
                range="'_NOTES_BACKUP'!A1",
                valueInputOption="RAW",
                body={"values": header}
            ).execute()

        # Read all rep tabs and collect notes that have any content
        meta      = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        tab_names = [s['properties']['title'] for s in meta['sheets']]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        backup_rows = []

        for tab in tab_names:
            if tab.startswith('_') or tab == 'MASTER':
                continue
            result = service.spreadsheets().values().get(
                spreadsheetId=SPREADSHEET_ID,
                range=f"'{tab}'!A1:Z5000"
            ).execute()
            rows = result.get('values', [])
            if len(rows) < 5:
                continue
            # Find header row (row 4, index 3)
            headers = None
            for r in rows:
                if 'CUSTOMER_NAME' in r:
                    headers = r
                    break
            if not headers:
                continue
            for row in rows[rows.index(headers)+1:]:
                row += [''] * (len(headers) - len(row))
                rec = dict(zip(headers, row))
                # Only backup rows that have at least one note field filled
                has_note = any(rec.get(c, '').strip() for c in NOTE_COLS)
                if has_note:
                    backup_rows.append([
                        timestamp,
                        rec.get('CUSTOMER_NAME', '').strip(),
                        rec.get('SALES_REP', '').strip(),
                    ] + [rec.get(c, '') for c in NOTE_COLS])

        if backup_rows:
            service.spreadsheets().values().append(
                spreadsheetId=SPREADSHEET_ID,
                range="'_NOTES_BACKUP'!A1",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": backup_rows}
            ).execute()
            print(f"  → Backed up {len(backup_rows)} note records to _NOTES_BACKUP.")
        else:
            print(f"  → No notes to back up yet.")

        # Format the backup tab header row
        backup_sheet_id = None
        meta2 = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        for s in meta2['sheets']:
            if s['properties']['title'] == '_NOTES_BACKUP':
                backup_sheet_id = s['properties']['sheetId']
                break
        if backup_sheet_id is not None:
            service.spreadsheets().batchUpdate(
                spreadsheetId=SPREADSHEET_ID,
                body={"requests": [
                    {"repeatCell": {
                        "range": {"sheetId": backup_sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                        "cell": {"userEnteredFormat": {
                            "backgroundColor": {"red": 0.122, "green": 0.22, "blue": 0.392},
                            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                           "bold": True, "fontSize": 9},
                        }},
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }},
                    {"updateSheetProperties": {
                        "properties": {"sheetId": backup_sheet_id, "gridProperties": {"frozenRowCount": 1}},
                        "fields": "gridProperties.frozenRowCount"
                    }},
                ]}
            ).execute()

    except Exception as e:
        print(f"  WARNING: Could not backup notes — {e}")


def push_rep_tab(service, rep, rep_df, year_list, all_tabs):
    """Write one rep's full account list to their Sheet tab with instructions and formatting."""
    if service is None:
        return
    try:
        h_tabs = get_all_tabs(service, HISTORY_SHEET_ID)
        ensure_tab_exists(service, rep, h_tabs, sheet_id=HISTORY_SHEET_ID)

        year_headers = [str(y) for y in year_list]
        headers = (
            ['CUSTOMER_NAME', 'CONTACT_NAME', 'CONTACT_TITLE', 'CONTACT_EMAIL',
             'CONTACT_PHONE', 'CONTACT_MOBILE',
             'SALES_REP', 'STATUS', 'LAST_SALE_DATE', 'DAYS_INACTIVE']
            + year_headers
            + ['TOTAL_2020_2026']
            + PRESERVE_COLS
        )
        num_cols = len(headers)

        data_rows = [headers]
        for _, r in rep_df.iterrows():
            row = [
                str(r.get('CUSTOMER_NAME', '')).strip(),
                r.get('CONTACT_NAME', ''),
                r.get('CONTACT_TITLE', ''),
                r.get('CONTACT_EMAIL', ''),
                r.get('CONTACT_PHONE', ''),
                r.get('CONTACT_MOBILE', ''),
                r.get('SALES_REP', ''),
                r.get('DECLINE_STATUS') or r.get('STATUS', ''),
                str(r.get('LAST_SALE_DATE', '')),
                str(r.get('DAYS_INACTIVE', '')),
            ]
            for y in year_list:
                col = f"Y{y}"
                val = r.get(col, 0)
                row.append(round(float(val), 2) if pd.notna(val) and val != '' else 0)
            row.append(round(float(r.get('TOTAL', 0)), 2) if pd.notna(r.get('TOTAL', 0)) else 0)
            for c in PRESERVE_COLS:
                row.append(r.get(c, ''))
            data_rows.append(row)

        # Instruction rows — pad to full width so they span all columns
        def pad(row_vals):
            return row_vals + [''] * (num_cols - len(row_vals))

        instr_rows = [
            pad([f"INTRANSIT TECHNOLOGIES — Account Review  |  Rep: {rep}  |  Generated: {datetime.now().strftime('%B %d, %Y')}  |  ⚠️ Fill in the LAST 5 yellow columns only — do not edit any other columns."]),
            pad(["Contact columns (Name/Title/Email/Phone) are read-only — pulled from CRM  |  REP_NOTE = Your comments  |  FOLLOW_UP_DATE = MM/DD/YYYY  |  REP_STATUS = Working / Quoted / Dead / Hold  |  REMOVE_FLAG = Type REQUEST to ask John to remove  |  JOHN_APPROVAL = John only — do not edit  |  HIDE_FROM_POOL = Admin only — do not edit"]),
            pad([]),
        ]
        all_rows = instr_rows + data_rows

        service.spreadsheets().values().clear(
            spreadsheetId=HISTORY_SHEET_ID,
            range=f"'{rep}'!A1:Z5000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=HISTORY_SHEET_ID,
            range=f"'{rep}'!A1",
            valueInputOption="RAW",
            body={"values": all_rows}
        ).execute()

        # Get sheet_id for formatting
        sheet_id = None
        meta = service.spreadsheets().get(spreadsheetId=HISTORY_SHEET_ID).execute()
        for s in meta['sheets']:
            if s['properties']['title'] == rep:
                sheet_id = s['properties']['sheetId']
                break

        if sheet_id is not None:
            note_start_col = len(headers) - len(PRESERVE_COLS)
            last_col_idx = num_cols - 1  # 0-based

            requests = [
                # Merge row 1 across all columns
                {"mergeCells": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                              "startColumnIndex": 0, "endColumnIndex": num_cols},
                    "mergeType": "MERGE_ALL"
                }},
                # Merge row 2 across all columns
                {"mergeCells": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2,
                              "startColumnIndex": 0, "endColumnIndex": num_cols},
                    "mergeType": "MERGE_ALL"
                }},
                # Row 1 style — dark blue, white bold, wrap
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                              "startColumnIndex": 0, "endColumnIndex": num_cols},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 0.122, "green": 0.22, "blue": 0.392},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                       "bold": True, "fontSize": 11, "fontFamily": "Arial"},
                        "wrapStrategy": "WRAP",
                        "verticalAlignment": "MIDDLE",
                        "horizontalAlignment": "LEFT",
                        "padding": {"top": 8, "bottom": 8, "left": 10, "right": 10}
                    }},
                    "fields": "userEnteredFormat(backgroundColor,textFormat,wrapStrategy,verticalAlignment,horizontalAlignment,padding)"
                }},
                # Row 1 height
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 0, "endIndex": 1},
                    "properties": {"pixelSize": 40},
                    "fields": "pixelSize"
                }},
                # Row 2 style — medium blue, white, wrap
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2,
                              "startColumnIndex": 0, "endColumnIndex": num_cols},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 0.184, "green": 0.329, "blue": 0.588},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                       "bold": False, "fontSize": 9, "fontFamily": "Arial"},
                        "wrapStrategy": "WRAP",
                        "verticalAlignment": "MIDDLE",
                        "horizontalAlignment": "LEFT",
                        "padding": {"top": 6, "bottom": 6, "left": 10, "right": 10}
                    }},
                    "fields": "userEnteredFormat(backgroundColor,textFormat,wrapStrategy,verticalAlignment,horizontalAlignment,padding)"
                }},
                # Row 2 height
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 1, "endIndex": 2},
                    "properties": {"pixelSize": 36},
                    "fields": "pixelSize"
                }},
                # Row 4 header — dark blue, white bold
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": 4,
                              "startColumnIndex": 0, "endColumnIndex": num_cols},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 0.122, "green": 0.22, "blue": 0.392},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                       "bold": True, "fontSize": 9, "fontFamily": "Arial"},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE"
                    }},
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)"
                }},
                # Yellow highlight on note columns (rows 5+)
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 4, "endRowIndex": 1000,
                              "startColumnIndex": note_start_col, "endColumnIndex": num_cols},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 1.0, "green": 0.95, "blue": 0.8},
                    }},
                    "fields": "userEnteredFormat(backgroundColor)"
                }},
                # Column A width — wide for customer names
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
                    "properties": {"pixelSize": 300},
                    "fields": "pixelSize"
                }},
                # Column B (SALES_REP) width
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2},
                    "properties": {"pixelSize": 90},
                    "fields": "pixelSize"
                }},
                # Column C (STATUS) width
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 2, "endIndex": 3},
                    "properties": {"pixelSize": 200},
                    "fields": "pixelSize"
                }},
                # Columns D-E (dates/days) width
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 3, "endIndex": 5},
                    "properties": {"pixelSize": 120},
                    "fields": "pixelSize"
                }},
                # Note columns width
                {"updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                              "startIndex": note_start_col, "endIndex": num_cols},
                    "properties": {"pixelSize": 160},
                    "fields": "pixelSize"
                }},
                # Freeze top 4 rows
                {"updateSheetProperties": {
                    "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 4}},
                    "fields": "gridProperties.frozenRowCount"
                }},
                # Dollar format on year columns and total (cols 5 to note_start_col)
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 4, "endRowIndex": 1000,
                              "startColumnIndex": 5, "endColumnIndex": note_start_col},
                    "cell": {"userEnteredFormat": {
                        "numberFormat": {"type": "CURRENCY", "pattern": "$#,##0.00"}
                    }},
                    "fields": "userEnteredFormat(numberFormat)"
                }},
                # Right-align dollar columns
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 4, "endRowIndex": 1000,
                              "startColumnIndex": 5, "endColumnIndex": note_start_col},
                    "cell": {"userEnteredFormat": {
                        "horizontalAlignment": "RIGHT"
                    }},
                    "fields": "userEnteredFormat(horizontalAlignment)"
                }},
                # Center-align DAYS_INACTIVE column (col 4)
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 4, "endRowIndex": 1000,
                              "startColumnIndex": 4, "endColumnIndex": 5},
                    "cell": {"userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }},
                    "fields": "userEnteredFormat(horizontalAlignment)"
                }},
            ]
            service.spreadsheets().batchUpdate(
                spreadsheetId=HISTORY_SHEET_ID,
                body={"requests": requests}
            ).execute()

        print(f"  → Pushed {len(rep_df)} accounts to Sheet tab: {rep} (History sheet)")
    except Exception as e:
        print(f"  WARNING: Could not push tab for {rep} — {e}")


def push_master_tab(service, summary, year_list, all_tabs):
    """Write the MASTER tab with all reps combined."""
    if service is None:
        return
    try:
        ensure_tab_exists(service, 'MASTER', all_tabs)
        year_headers = [str(y) for y in year_list]
        headers = (
            ['CUSTOMER_NAME', 'SALES_REP', 'STATUS', 'LAST_SALE_DATE', 'DAYS_INACTIVE']
            + year_headers
            + ['TOTAL_2020_2026']
            + PRESERVE_COLS
        )
        rows = [headers]
        for _, r in summary.iterrows():
            row = [
                r.get('CUSTOMER_NAME', ''),
                r.get('SALES_REP', ''),
                r.get('DECLINE_STATUS') or r.get('STATUS', ''),
                str(r.get('LAST_SALE_DATE', '')),
                str(r.get('DAYS_INACTIVE', '')),
            ]
            for y in year_list:
                col = f"Y{y}"
                val = r.get(col, 0)
                row.append(round(float(val), 2) if pd.notna(val) and val != '' else 0)
            row.append(round(float(r.get('TOTAL', 0)), 2) if pd.notna(r.get('TOTAL', 0)) else 0)
            for c in PRESERVE_COLS:
                row.append(r.get(c, ''))
            rows.append(row)

        service.spreadsheets().values().clear(
            spreadsheetId=SPREADSHEET_ID,
            range="'MASTER'!A1:Z5000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range="'MASTER'!A1",
            valueInputOption="RAW",
            body={"values": rows}
        ).execute()
        print(f"  → Pushed {len(summary)} accounts to MASTER tab.")
    except Exception as e:
        print(f"  WARNING: Could not push MASTER tab — {e}")


def log_run_to_sheet(service, stats, all_tabs):
    """Append a run record to the _LOG tab (History sheet)."""
    if service is None:
        return
    try:
        sid = HISTORY_SHEET_ID
        h_tabs = get_all_tabs(service, sid)
        ensure_tab_exists(service, '_LOG', h_tabs, sheet_id=sid)
        # Write header if empty
        existing = service.spreadsheets().values().get(
            spreadsheetId=sid, range="'_LOG'!A1:A1"
        ).execute()
        if not existing.get('values'):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range="'_LOG'!A1",
                valueInputOption="RAW",
                body={"values": [['RUN_DATE', 'TOTAL_ACCOUNTS', 'INACTIVE', 'AT_RISK', 'ACTIVE', 'NOTE']]}
            ).execute()
        # Append log row
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            stats['total'], stats['inactive'], stats['at_risk'], stats['active'],
            stats.get('note', '')
        ]
        service.spreadsheets().values().append(
            spreadsheetId=sid,
            range="'_LOG'!A1",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": [row]}
        ).execute()
        print("  → Run logged to _LOG tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not write to _LOG — {e}")



def push_orders_history_tab(service, orders_hist_df, all_tabs):
    """Push full orders history (dated) to _ORDERS_HISTORY tab in HISTORY sheet."""
    if service is None or len(orders_hist_df) == 0:
        return
    sid = HISTORY_SHEET_ID
    h_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_ORDERS_HISTORY', h_tabs, row_count=len(orders_hist_df), sheet_id=sid)
        resize_tab(service, '_ORDERS_HISTORY', len(orders_hist_df) + 1, sheet_id=sid)
        headers = ['ORDER_NUMBER', 'CUSTOMER_NAME', 'SALES_REP', 'ORDER_DATE',
                   'PART_NUMBER', 'DESCRIPTION', 'EXTENDED_PRICE', 'COST', 'GP', 'CONDITION']
        rows = [headers]
        for _, r in orders_hist_df.iterrows():
            rows.append([str(r.get(h, '')) for h in headers])
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_ORDERS_HISTORY'!A1:Z200000"
        ).execute()
        for i in range(0, len(rows), 10000):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_ORDERS_HISTORY'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+10000]}
            ).execute()
        _format_header_tab(service, '_ORDERS_HISTORY', sid=sid)
        print(f"  -> Pushed {len(rows)-1} order history records to _ORDERS_HISTORY tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _ORDERS_HISTORY tab -- {e}")

def push_open_orders_tab(service, orders_df, all_tabs):
    """Push open sales orders to _OPEN_ORDERS tab."""
    if service is None or len(orders_df) == 0:
        return
    try:
        h_tabs = get_all_tabs(service, HISTORY_SHEET_ID)
        ensure_tab_exists(service, '_OPEN_ORDERS', h_tabs, row_count=len(orders_df), sheet_id=HISTORY_SHEET_ID)
        resize_tab(service, '_OPEN_ORDERS', len(orders_df) + 1, sheet_id=HISTORY_SHEET_ID)
        headers = ['ORDER_NUMBER','CUSTOMER_ID','CUSTOMER_NAME','SALES_REP',
                   'ORDER_DATE','PART_NUMBER','DESCRIPTION','QTY_ORDERED',
                   'QTY_SHIPPED','QTY_NEEDED','PRICE','EXTENDED_PRICE',
                   'COST','CONDITION']
        rows = [headers]
        for _, r in orders_df.iterrows():
            rows.append([str(r.get(h,'')) for h in headers])
        service.spreadsheets().values().clear(
            spreadsheetId=HISTORY_SHEET_ID, range="'_OPEN_ORDERS'!A1:Z200000"
        ).execute()
        for i in range(0, len(rows), 10000):
            service.spreadsheets().values().update(
                spreadsheetId=HISTORY_SHEET_ID,
                range=f"'_OPEN_ORDERS'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+10000]}
            ).execute()
        _format_header_tab(service, '_OPEN_ORDERS', sid=HISTORY_SHEET_ID)
        print(f"  → Pushed {len(rows)-1} open order lines to _OPEN_ORDERS tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _OPEN_ORDERS tab — {e}")


def push_gp_tab(service, gp_df, all_tabs):
    """Push GP data to _GP tab in HISTORY sheet."""
    if service is None or len(gp_df) == 0:
        return
    sid = HISTORY_SHEET_ID
    h_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_GP', h_tabs, row_count=len(gp_df), sheet_id=sid)
        resize_tab(service, '_GP', len(gp_df) + 1, sheet_id=sid)
        headers = ['INVOICE_NUMBER','INVOICE_DATE','SALES_REP','CUSTOMER_NAME',
                   'EXTENDED_PRICE','GP','PART_NUMBER','INVOICE_MONTH','INVOICE_YEAR','BUYER_NAME']
        rows = [headers]
        for _, r in gp_df.iterrows():
            rows.append([str(r.get(h,'')) for h in headers])
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_GP'!A1:Z200000"
        ).execute()
        for i in range(0, len(rows), 10000):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_GP'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+10000]}
            ).execute()
        _format_header_tab(service, '_GP', sid=sid)
        print(f"  → Pushed {len(rows)-1} GP records to _GP tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _GP tab — {e}")


def resize_tab(service, tab_name, needed_rows, needed_cols=20, sheet_id=None):
    """Resize an existing tab to fit the data."""
    sid = sheet_id or SPREADSHEET_ID
    try:
        meta = service.spreadsheets().get(spreadsheetId=sid).execute()
        for s in meta['sheets']:
            if s['properties']['title'] == tab_name:
                cur_rows = s['properties']['gridProperties']['rowCount']
                cur_cols = s['properties']['gridProperties']['columnCount']
                if cur_rows < needed_rows or cur_cols < needed_cols:
                    service.spreadsheets().batchUpdate(
                        spreadsheetId=sid,
                        body={"requests": [{"updateSheetProperties": {
                            "properties": {
                                "sheetId": s['properties']['sheetId'],
                                "gridProperties": {
                                    "rowCount": max(cur_rows, needed_rows + 100),
                                    "columnCount": max(cur_cols, needed_cols)
                                }
                            },
                            "fields": "gridProperties.rowCount,gridProperties.columnCount"
                        }}]}
                    ).execute()
                break
    except Exception as e:
        print(f"    (Could not resize {tab_name}: {e})")


def push_reqs_tab(service, reqs_df, all_tabs):
    """Push RFQ history to _REQS tab in ACTIVITY sheet."""
    if service is None or len(reqs_df) == 0:
        return
    sid = ACTIVITY_SHEET_ID
    a_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_REQS', a_tabs, row_count=len(reqs_df), sheet_id=sid)
        resize_tab(service, '_REQS', len(reqs_df) + 1, sheet_id=sid)
        headers = ['REQ_ID','CONTACT_ID','CUSTOMER_ID','CUSTOMER_NAME','SALES_REP',
                   'DATE_ENTERED','PART_NUMBER','MANUFACTURER','QTY',
                   'TARGET_PRICE','DESCRIPTION','CONDITION',
                   'CLOSED','DATE_CLOSED','SELLER']
        rows = [headers]
        for _, r in reqs_df.iterrows():
            rows.append([str(r.get(h,'')) for h in headers])

        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_REQS'!A1:Z200000"
        ).execute()
        for i in range(0, len(rows), 10000):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_REQS'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+10000]}
            ).execute()

        _format_header_tab(service, '_REQS', sid=sid)
        print(f"  → Pushed {len(rows)-1} RFQ records to _REQS tab (Activity sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _REQS tab — {e}")


def push_quotes_tab(service, quotes_df, all_tabs):
    """Push quote history to _QUOTES tab in ACTIVITY sheet."""
    if service is None or len(quotes_df) == 0:
        return
    sid = ACTIVITY_SHEET_ID
    a_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_QUOTES', a_tabs, row_count=len(quotes_df), sheet_id=sid)
        resize_tab(service, '_QUOTES', len(quotes_df) + 1, sheet_id=sid)

        headers = ['QUOTE_ID','CONTACT_ID','CUSTOMER_ID','CUSTOMER_NAME','SALES_REP',
                   'QUOTE_DATE','CLOSED','ENTERED_BY','PART_NUMBER','MANUFACTURER',
                   'QTY','PRICE','DESCRIPTION','CONDITION','LINE_CLOSED']
        rows = [headers]
        for _, r in quotes_df.iterrows():
            rows.append([str(r.get(h,'')) for h in headers])

        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_QUOTES'!A1:Z200000"
        ).execute()
        for i in range(0, len(rows), 10000):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_QUOTES'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+10000]}
            ).execute()

        _format_header_tab(service, '_QUOTES', sid=sid)
        print(f"  → Pushed {len(rows)-1} quote records to _QUOTES tab (Activity sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _QUOTES tab — {e}")


def push_customer_stats_tab(service, cust_stats_df, all_tabs):
    """Push customer stats to _CUSTOMER_STATS tab in HISTORY sheet."""
    if service is None or len(cust_stats_df) == 0:
        return
    sid = HISTORY_SHEET_ID
    h_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_CUSTOMER_STATS', h_tabs, sheet_id=sid)
        headers = ['CUSTOMER_ID','CUSTOMER_NAME','SALES_REP','ESTABLISHED',
                   'NUM_OF_BUYERS','NUM_OF_EMP','REVENUE',
                   'OPEN_REQS','OPEN_QUOTES','OPEN_ORDERS',
                   'TOTAL_REQS','TOTAL_QUOTES','TOTAL_INVOICES',
                   'TOTAL_ORDERS','TOTAL_RETURNS','LY_SALES',
                   'AVERAGE_PAY','CURRENT_BALANCE','CREDIT_LIMIT','CREDIT_HOLD',
                   'LAST_REQ','LAST_QUOTE','LAST_ORDER','LAST_INVOICE',
                   'SEGMENT','ACCOUNT_CLASS','ACCOUNT_TYPE',
                   'LAST_ACTIVITY','LAST_CONTACTED']
        rows = [headers]
        for _, r in cust_stats_df.iterrows():
            rows.append([str(r.get(h,'')) for h in headers])

        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_CUSTOMER_STATS'!A1:Z10000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sid,
            range="'_CUSTOMER_STATS'!A1",
            valueInputOption="RAW",
            body={"values": rows}
        ).execute()

        _format_header_tab(service, '_CUSTOMER_STATS', sid=sid)
        print(f"  → Pushed {len(rows)-1} customer stat records to _CUSTOMER_STATS tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _CUSTOMER_STATS tab — {e}")


def push_customer_directory_tab(service, conn, all_tabs):
    """Push all CRM customers (by active rep) to _CUSTOMER_DIRECTORY tab in History sheet.
    Used by the app's CRM Discovery panel and AI Picker on the Attack Plan page."""
    if service is None:
        return
    sid = HISTORY_SHEET_ID
    h_tabs = get_all_tabs(service, sid)
    print(f"  → _CUSTOMER_DIRECTORY: History sheet tabs found: {len(h_tabs)}")
    # Open a fresh connection — the main conn may already be closed by this point
    _conn = None
    try:
        _conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=localhost;"
            "DATABASE=CCCRM;"
            "Trusted_Connection=yes;"
        )
        ensure_tab_exists(service, '_CUSTOMER_DIRECTORY', h_tabs, sheet_id=sid)
        print(f"  → _CUSTOMER_DIRECTORY: Tab ensured.")

        dir_sql = """
            WITH owned_accounts AS (
                SELECT arv.ACCOUNT AS CUSTOMER_ID,
                       RTRIM(e.LOGIN_ID) AS SALES_REP,
                       ROW_NUMBER() OVER (PARTITION BY arv.ACCOUNT ORDER BY e.LOGIN_ID) AS rn
                FROM dbo.ACCOUNT_REP_VIEW arv
                JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
            )
            SELECT
                c.ID            AS CUSTOMER_ID,
                c.NAME          AS NAME,
                oa.SALES_REP    AS USERNAME,
                c.CITY          AS CITY,
                c.STATE         AS STATE,
                c.LAST_ACTIVITY AS LAST_ACTIVITY,
                ISNULL(ytd.YTD_REVENUE, 0)      AS YTD_REVENUE,
                ISNULL(life.LIFETIME_REVENUE, 0) AS LIFETIME_REVENUE,
                life.FIRST_SALE_DATE,
                life.LAST_SALE_DATE,
                ISNULL(life.TOTAL_ORDERS, 0)    AS TOTAL_ORDERS
            FROM dbo.CUSTOMER c
            JOIN owned_accounts oa ON oa.CUSTOMER_ID = c.ID AND oa.rn = 1
            -- YTD revenue (all orders for this customer, any entry clerk)
            LEFT JOIN (
                SELECT o.CUSTOMER_ID, SUM(i.INVOICE_TOTAL) AS YTD_REVENUE
                FROM dbo.INVCHEA i
                JOIN dbo.ORDERHEA o ON o.ORDER_NUMBER = i.ORDER_NUMBER
                WHERE YEAR(i.INVOICE_DATE) = YEAR(GETDATE())
                GROUP BY o.CUSTOMER_ID
            ) ytd ON ytd.CUSTOMER_ID = c.ID
            -- Lifetime revenue (all orders for this customer, any entry clerk)
            LEFT JOIN (
                SELECT o.CUSTOMER_ID,
                       SUM(i.INVOICE_TOTAL)        AS LIFETIME_REVENUE,
                       MIN(i.INVOICE_DATE)          AS FIRST_SALE_DATE,
                       MAX(i.INVOICE_DATE)          AS LAST_SALE_DATE,
                       COUNT(DISTINCT i.INVOICE_NUMBER) AS TOTAL_ORDERS
                FROM dbo.INVCHEA i
                JOIN dbo.ORDERHEA o ON o.ORDER_NUMBER = i.ORDER_NUMBER
                GROUP BY o.CUSTOMER_ID
            ) life ON life.CUSTOMER_ID = c.ID
            WHERE c.NAME IS NOT NULL
              AND LEN(LTRIM(RTRIM(c.NAME))) > 0
            ORDER BY oa.SALES_REP, c.NAME
        """
        print(f"  → _CUSTOMER_DIRECTORY: Running SQL query...")
        dir_df = pd.read_sql(dir_sql, _conn)
        print(f"  → Pulled {len(dir_df):,} customers for directory.")

        if len(dir_df) == 0:
            print(f"  WARNING: SQL returned 0 rows — check USERNAME values in dbo.CUSTOMER")
            return

        headers = ['CUSTOMER_ID', 'NAME', 'USERNAME', 'CITY', 'STATE', 'LAST_ACTIVITY',
                   'YTD_REVENUE', 'LIFETIME_REVENUE', 'FIRST_SALE_DATE', 'LAST_SALE_DATE', 'TOTAL_ORDERS']
        rows = [headers]
        for _, r in dir_df.iterrows():
            rows.append([
                str(r.get('CUSTOMER_ID', '') or ''),
                str(r.get('NAME', '') or '').strip(),
                str(r.get('USERNAME', '') or '').strip(),
                str(r.get('CITY', '') or '').strip(),
                str(r.get('STATE', '') or '').strip(),
                str(r.get('LAST_ACTIVITY', '') or '').strip(),
                round(float(r.get('YTD_REVENUE', 0) or 0), 2),
                round(float(r.get('LIFETIME_REVENUE', 0) or 0), 2),
                str(r.get('FIRST_SALE_DATE', '') or '')[:10],
                str(r.get('LAST_SALE_DATE', '') or '')[:10],
                int(r.get('TOTAL_ORDERS', 0) or 0),
            ])

        resize_tab(service, '_CUSTOMER_DIRECTORY', len(rows) + 100, sheet_id=sid)
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_CUSTOMER_DIRECTORY'!A1:K50000"
        ).execute()
        chunk_size = 10000
        for i in range(0, len(rows), chunk_size):
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_CUSTOMER_DIRECTORY'!A{i+1}",
                valueInputOption="RAW",
                body={"values": rows[i:i+chunk_size]}
            ).execute()
            print(f"  → _CUSTOMER_DIRECTORY: Wrote rows {i+1}–{min(i+chunk_size, len(rows))}")

        _format_header_tab(service, '_CUSTOMER_DIRECTORY', sid=sid)
        print(f"  → Pushed {len(rows)-1} accounts to _CUSTOMER_DIRECTORY tab (History sheet).")
    except Exception as e:
        print(f"  ERROR: Could not push _CUSTOMER_DIRECTORY tab — {e}")
        traceback.print_exc()
    finally:
        if _conn:
            try:
                _conn.close()
            except Exception:
                pass


def _format_header_tab(service, tab_name, sid=None):
    """Apply standard dark blue header formatting to a tab."""
    sid = sid or SPREADSHEET_ID
    try:
        meta = service.spreadsheets().get(spreadsheetId=sid).execute()
        sheet_id = None
        for s in meta['sheets']:
            if s['properties']['title'] == tab_name:
                sheet_id = s['properties']['sheetId']
                break
        if sheet_id is not None:
            service.spreadsheets().batchUpdate(
                spreadsheetId=sid,
                body={"requests": [
                    {"repeatCell": {
                        "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                        "cell": {"userEnteredFormat": {
                            "backgroundColor": {"red": 0.122, "green": 0.22, "blue": 0.392},
                            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                           "bold": True, "fontSize": 9},
                        }},
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }},
                    {"updateSheetProperties": {
                        "properties": {"sheetId": sheet_id,
                                       "gridProperties": {"frozenRowCount": 1}},
                        "fields": "gridProperties.frozenRowCount"
                    }},
                ]}
            ).execute()
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
#  NEW DATA TABS — CONTACTS / INVOICE HISTORY / COLLECTIONS
# ══════════════════════════════════════════════════════════════════════════════

def push_contacts_tab(service, contacts_all_df, all_tabs):
    """Push ALL contacts for every account to _CONTACTS tab.
    Flags AP/Finance/Collections contacts separately from purchasing contacts."""
    if service is None:
        return
    try:
        sid = HISTORY_SHEET_ID
        h_tabs = get_all_tabs(service, sid)
        ensure_tab_exists(service, '_CONTACTS', h_tabs, sheet_id=sid)

        AP_KEYWORDS = ['payable', 'accounts pay', 'ap ', 'a/p', 'finance',
                       'controller', 'treasury', 'billing', 'credit', 'collection']

        headers = ['CUSTOMER_ID', 'CUSTOMER_NAME', 'SALES_REP',
                   'CONTACT_NAME', 'TITLE', 'EMAIL', 'PHONE', 'MOBILE',
                   'CONTACT_TYPE', 'HAS_EMAIL']
        rows = [headers]
        for _, r in contacts_all_df.iterrows():
            title_lower = str(r.get('TITLE', '')).lower()
            contact_type = 'AP/Collections' if any(k in title_lower for k in AP_KEYWORDS) else 'Purchasing'
            has_email = 'YES' if str(r.get('EMAIL', '')).strip() else 'NO'
            rows.append([
                str(r.get('CUSTOMER_ID', '')),
                str(r.get('CUSTOMER_NAME', '')),
                str(r.get('SALES_REP', '')),
                str(r.get('CONTACT_NAME', '')),
                str(r.get('TITLE', '')),
                str(r.get('EMAIL', '')),
                str(r.get('PHONE', '')),
                str(r.get('MOBILE', '')),
                contact_type,
                has_email,
            ])

        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_CONTACTS'!A1:Z50000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sid,
            range="'_CONTACTS'!A1",
            valueInputOption="RAW",
            body={"values": rows}
        ).execute()

        # Format header
        sheet_id = None
        meta = service.spreadsheets().get(spreadsheetId=sid).execute()
        for s in meta['sheets']:
            if s['properties']['title'] == '_CONTACTS':
                sheet_id = s['properties']['sheetId']
                break
        if sheet_id is not None:
            service.spreadsheets().batchUpdate(
                spreadsheetId=sid,
                body={"requests": [
                    {"repeatCell": {
                        "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                        "cell": {"userEnteredFormat": {
                            "backgroundColor": {"red": 0.122, "green": 0.22, "blue": 0.392},
                            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                           "bold": True, "fontSize": 9},
                        }},
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }},
                    {"updateSheetProperties": {
                        "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                        "fields": "gridProperties.frozenRowCount"
                    }},
                ]}
            ).execute()

        ap_count = sum(1 for r in rows[1:] if r[8] == 'AP/Collections')
        print(f"  → Pushed {len(rows)-1} contacts to _CONTACTS tab (History sheet) "
              f"({ap_count} AP/Collections contacts flagged).")
    except Exception as e:
        print(f"  WARNING: Could not push _CONTACTS tab — {e}")


def push_invoice_history_tab(service, invoice_history_df, all_tabs):
    """Push recent invoice line items to _INVOICE_HISTORY tab in HISTORY sheet."""
    if service is None:
        return
    sid = HISTORY_SHEET_ID
    h_tabs = get_all_tabs(service, sid)
    try:
        ensure_tab_exists(service, '_INVOICE_HISTORY', h_tabs, sheet_id=sid)

        headers = ['CUSTOMER_ID', 'CUSTOMER_NAME', 'SALES_REP',
                   'INVOICE_NUMBER', 'INVOICE_DATE', 'INVOICE_TOTAL',
                   'PART_NUMBER', 'DESCRIPTION', 'QTY', 'UNIT_PRICE', 'LINE_TOTAL']
        rows = [headers]
        for _, r in invoice_history_df.iterrows():
            rows.append([
                str(r.get('CUSTOMER_ID', '')),
                str(r.get('CUSTOMER_NAME', '')),
                str(r.get('SALES_REP', '')),
                str(r.get('INVOICE_NUMBER', '')),
                str(r.get('INVOICE_DATE', '')),
                round(float(r.get('INVOICE_TOTAL', 0) or 0), 2),
                str(r.get('PART_NUMBER', '')),
                str(r.get('DESCRIPTION', '')),
                str(r.get('QTY', '')),
                round(float(r.get('UNIT_PRICE', 0) or 0), 2),
                round(float(r.get('LINE_TOTAL', 0) or 0), 2),
            ])

        resize_tab(service, '_INVOICE_HISTORY', len(rows) + 10, sheet_id=sid)
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_INVOICE_HISTORY'!A1:Z200000"
        ).execute()

        chunk_size = 10000
        for i in range(0, len(rows), chunk_size):
            chunk = rows[i:i+chunk_size]
            service.spreadsheets().values().update(
                spreadsheetId=sid,
                range=f"'_INVOICE_HISTORY'!A{i+1}",
                valueInputOption="RAW",
                body={"values": chunk}
            ).execute()

        _format_header_tab(service, '_INVOICE_HISTORY', sid=sid)
        print(f"  → Pushed {len(rows)-1} invoice line items to _INVOICE_HISTORY tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _INVOICE_HISTORY tab — {e}")



def push_collections_invoices_tab(service, all_tabs):
    """Push open invoice details for collections accounts to _COLLECTIONS_INVOICES tab."""
    if service is None:
        return
    try:
        sid = HISTORY_SHEET_ID
        h_tabs = get_all_tabs(service, sid)
        ensure_tab_exists(service, '_COLLECTIONS_INVOICES', h_tabs, sheet_id=sid)
        
        # Pull open invoices with balance > 0
        import pyodbc as _pyodbc
        conn2 = _pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=localhost;DATABASE=CCCRM;Trusted_Connection=yes;"
        )
        import pandas as _pd
        sql = """
            SELECT
                i.INVOICE_NUMBER,
                i.INVOICE_DATE,
                i.INVOICE_TOTAL,
                i.BALANCE                AS INVOICE_BALANCE,
                i.TRACKING_NUMBER,
                i.SHIP_VIA,
                c.NAME                   AS CUSTOMER_NAME,
                (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
                 JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
                 WHERE arv2.ACCOUNT = i.CUSTOMER_ID
                   AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
                 ORDER BY e2.LOGIN_ID) AS SALES_REP,
                c.TERMS                  AS PAYMENT_TERMS,
                o.ORDER_NUMBER           AS ORDER_NUMBER,
                i.CUSTOMER_ID,
                DATEDIFF(day, i.INVOICE_DATE, GETDATE()) AS AGING_DAYS
            FROM dbo.INVCHEA i
            JOIN dbo.ORDERHEA o ON o.ORDER_NUMBER = i.ORDER_NUMBER
            JOIN dbo.CUSTOMER c ON c.ID = i.CUSTOMER_ID
            WHERE i.BALANCE > 0
              AND i.INVOICE_DATE >= '2020-01-01'
              AND i.CUSTOMER_ID IN (
                  SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
                  JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                  WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
              )
            ORDER BY c.NAME, i.INVOICE_DATE DESC
        """
        inv_df = _pd.read_sql(sql, conn2)
        conn2.close()
        
        if len(inv_df) == 0:
            print("  → No open invoice balances found for _COLLECTIONS_INVOICES.")
            return
        
        headers = ['CUSTOMER_NAME','SALES_REP','INVOICE_NUMBER','INVOICE_DATE',
                   'INVOICE_TOTAL','INVOICE_BALANCE','PAYMENT_TERMS',
                   'ORDER_NUMBER','TRACKING_NUMBER','SHIP_VIA',
                   'CUSTOMER_ID','AGING_DAYS']
        rows = [headers]
        for _, r in inv_df.iterrows():
            rows.append([
                str(r.get('CUSTOMER_NAME','')).strip(),
                str(r.get('SALES_REP','')).strip(),
                str(r.get('INVOICE_NUMBER','')),
                str(r.get('INVOICE_DATE',''))[:10] if r.get('INVOICE_DATE') else '',
                round(float(r.get('INVOICE_TOTAL',0) or 0), 2),
                round(float(r.get('INVOICE_BALANCE',0) or 0), 2),
                str(r.get('PAYMENT_TERMS','')).strip(),
                str(r.get('ORDER_NUMBER','')).strip(),
                str(r.get('TRACKING_NUMBER','')).strip(),
                str(r.get('SHIP_VIA','')).strip(),
                str(r.get('CUSTOMER_ID','')),
                int(r.get('AGING_DAYS',0) or 0),
            ])
        
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_COLLECTIONS_INVOICES'!A1:Z50000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sid,
            range="'_COLLECTIONS_INVOICES'!A1",
            valueInputOption="RAW",
            body={"values": rows}
        ).execute()
        _format_header_tab(service, '_COLLECTIONS_INVOICES', sid=sid)
        print(f"  -> Pushed {len(rows)-1} open invoice records to _COLLECTIONS_INVOICES tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _COLLECTIONS_INVOICES tab -- {e}")

def push_collections_tab(service, collections_df, all_tabs):
    """Push collections data — balances, payment history, collection notes."""
    if service is None:
        return
    try:
        sid = HISTORY_SHEET_ID
        h_tabs = get_all_tabs(service, sid)
        ensure_tab_exists(service, '_COLLECTIONS', h_tabs, sheet_id=sid)

        headers = ['CUSTOMER_ID', 'CUSTOMER_NAME', 'SALES_REP',
                   'CURRENT_BALANCE', 'CREDIT_LIMIT', 'AVG_DAYS_TO_PAY',
                   'PAYMENT_TERMS', 'DAYS_PAST_DUE',
                   'LAST_PAYMENT_DATE', 'LAST_PAYMENT_AMOUNT',
                   'COLLECTION_NOTE', 'NOTE_DATE', 'NOTE_BY',
                   'RISK_FLAG', 'APP_HOLD']
        rows = [headers]
        for _, r in collections_df.iterrows():
            balance = float(r.get('CURRENT_BALANCE', 0) or 0)

            # Calculate days past due from terms + last invoice date
            terms_str = str(r.get('PAYMENT_TERMS', '') or '').strip().upper()
            terms_days = 30  # default
            import re as _re
            if any(t in terms_str for t in ['COD','CASH','CIA','PREPAID','T/T','TT','WIRE','TELEGRAPHIC','ADVANCE','CBD']):
                terms_days = 0
            else:
                m = _re.search(r'(\d+)', terms_str)
                if m:
                    terms_days = int(m.group(1))

            last_inv = r.get('LAST_INVOICE_DATE')
            days_past = 0
            if last_inv and str(last_inv) not in ('nan','NaT','None',''):
                try:
                    from datetime import date as _date
                    if hasattr(last_inv, 'date'):
                        inv_date = last_inv.date()
                    else:
                        inv_date = pd.to_datetime(str(last_inv)).date()
                    # Ignore bad/default dates before year 2000
                    if inv_date.year >= 2000:
                        due_date = inv_date + __import__('datetime').timedelta(days=terms_days)
                        days_past = max(0, (_date.today() - due_date).days)
                except Exception:
                    days_past = 0

            avg_days = float(r.get('AVG_DAYS_OVER', 0) or 0)

            # Risk flag logic based on terms
            if days_past > 30:
                risk = '🔴 HIGH RISK'
            elif days_past > 0:
                risk = '🟠 WATCH'
            else:
                risk = ''

            def clean(v):
                """Convert NaN/NaT/None to empty string."""
                import math
                if v is None: return ''
                s = str(v)
                if s in ('nan','NaN','NaT','None',''): return ''
                try:
                    if math.isnan(float(s)): return ''
                except: pass
                return s

            rows.append([
                clean(r.get('CUSTOMER_ID', '')),
                clean(r.get('CUSTOMER_NAME', '')),
                clean(r.get('SALES_REP', '')),
                round(balance, 2),
                round(float(r.get('CREDIT_LIMIT', 0) or 0), 2),
                round(avg_days, 1),
                clean(r.get('PAYMENT_TERMS', '')),
                days_past,
                clean(r.get('LAST_PAYMENT_DATE', '')),
                round(float(str(r.get('LAST_PAYMENT_AMOUNT', 0) or 0).replace('nan','0')), 2),
                clean(r.get('COLLECTION_NOTE', '')),
                clean(r.get('NOTE_DATE', '')),
                clean(r.get('NOTE_BY', '')),
                risk,
                '',  # APP_HOLD — managed in app only
            ])

        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_COLLECTIONS'!A1:Z10000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sid,
            range="'_COLLECTIONS'!A1",
            valueInputOption="RAW",
            body={"values": rows}
        ).execute()
        _format_header_tab(service, '_COLLECTIONS', sid=sid)

        high_risk = sum(1 for r in rows[1:] if '🔴' in str(r[11]))
        watch = sum(1 for r in rows[1:] if '🟠' in str(r[11]))
        print("  → Pushing unowned accounts (no rep in CRM)...")
        push_unowned_accounts_tab(service, h_tabs, collections_df, sid)
        print(f"  → Pushed {len(rows)-1} accounts to _COLLECTIONS tab "
              f"({high_risk} high risk, {watch} watch).")
    except Exception as e:
        print(f"  WARNING: Could not push _COLLECTIONS tab — {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def hdr_cell(ws, row, col, value, bg=C_DARK_BLUE, fg=C_WHITE, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, value, bold=False, color=C_BLACK, bg=None, num_fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=9, name="Arial")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    c.alignment = Alignment(horizontal=align, vertical="center")
    return c

def status_color(status):
    s = status or ''
    if '🔴' in s or 'High Value' in s:
        return ("C00000", "FFFFFF")
    if '🟠' in s or 'Declining' in s:
        return ("E26B0A", "FFFFFF")
    if '⚫' in s or 'Dormant' in s:
        return ("595959", "FFFFFF")
    if '🟡' in s or 'Monitor' in s:
        return ("BF9000", "FFFFFF")
    return (None, C_BLACK)


def build_rep_excel(rep, rep_df, output_path, year_list):
    """Build a full AccountReview-style Excel file for one rep."""
    wb = Workbook()

    # ── Sheet 1: Instructions ─────────────────────────────────────────────────
    ws_instr = wb.active
    ws_instr.title = "📋 Instructions"
    ws_instr.sheet_view.showGridLines = False

    inactive_df  = rep_df[rep_df['IS_INACTIVE'] == True]
    declining_df = rep_df[rep_df['IS_DECLINING'] == True]
    total_lost   = rep_df['TOTAL'].sum()

    ws_instr.merge_cells("A1:N1")
    t = ws_instr["A1"]
    t.value = f"INTRANSIT TECHNOLOGIES  —  Account Review Report"
    t.font  = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    t.fill  = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_instr.row_dimensions[1].height = 32

    ws_instr.merge_cells("A2:N2")
    sub = ws_instr["A2"]
    sub.value = (
        f"Sales Rep: {rep}  |  Prepared: {TODAY.strftime('%B %d, %Y')}  |  "
        f"Inactive Accounts: {len(inactive_df)}  |  Declining Accounts: {len(declining_df)}  |  "
        f"Total Lost Revenue: ${total_lost:,.0f}"
    )
    sub.font  = Font(bold=False, color=C_WHITE, size=10, name="Arial")
    sub.fill  = PatternFill("solid", fgColor=C_MED_BLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_instr.row_dimensions[2].height = 20

    instructions = [
        ("", ""),
        ("PURPOSE OF THIS REPORT", ""),
        ("", "This report has been prepared to help you identify accounts that need immediate attention. It is divided into two sections:"),
        ("", "     •  INACTIVE ACCOUNTS — Customers who have not placed an order in the past 6 months. These relationships need to be re-engaged."),
        ("", "     •  ACTIVE BUT DECLINING — Customers who are still buying but at a significantly lower pace than last year. These are at risk of going fully inactive."),
        ("", ""),
        ("YOUR ACTION PLAN", ""),
        ("", "Step 1 — Review each account on the INACTIVE tab. For every account, ask yourself: When did I last speak with this customer? Do I know why they stopped buying?"),
        ("", "Step 2 — Use the Notes / Action column to record your planned next step. Examples: 'Call scheduled for 5/10', 'Sent quote 5/6', 'Customer moved supplier — confirm'."),
        ("", "Step 3 — Review the DECLINING tab. These customers are still active but the trend is heading in the wrong direction. A proactive call now is much easier than recovering a lost account later."),
        ("", "Step 4 — Return the completed file to Karen with your notes filled in by the agreed deadline. Karen will review with you before any account reassignments are made."),
        ("", ""),
        ("HOW TO USE THE NOTE COLUMNS", ""),
        ("", "There are 4 columns at the end of each row for you to fill in. Your notes are saved every time the report runs — they will never be wiped out."),
        ("", ""),
        ("REP_NOTE", "Your comments about this account. Examples: 'Called 5/6, left voicemail', 'Sent bearing quote, waiting on response', 'Customer said they switched to local supplier'. Write anything useful here."),
        ("", ""),
        ("FOLLOW_UP_DATE", "The date you plan to follow up with this account. Format: MM/DD/YYYY — example: 05/15/2026. This date carries forward every report run so you never lose it."),
        ("", ""),
        ("REP_STATUS", "A short label to track where things stand. Suggested values:"),
        ("", "     •  Working    — You are actively pursuing this account"),
        ("", "     •  Quoted     — You have sent a quote and are waiting for a response"),
        ("", "     •  Dead       — This account is lost for good, not worth pursuing further"),
        ("", "     •  Hold       — Pause outreach for now, revisit later"),
        ("", ""),
        ("REMOVE_FLAG", "Use this ONLY to request that an account be removed from future reports. Type REQUEST and explain why in the REP_NOTE column."),
        ("", "     ⚠️  Do NOT type YES in this column. Removals must be approved by Karen first."),
        ("", "     ⚠️  Karen will review your request and set the final approval. Until she approves, the account stays in the report."),
        ("", "     Example: Type 'REQUEST' in REMOVE_FLAG and 'Company closed permanently' in REP_NOTE."),
        ("", ""),
        ("IMPORTANT", ""),
        ("", "No accounts will be reassigned without a conversation with you and management first. This report is a tool to help prioritize your efforts — not a performance evaluation."),
        ("", "If you believe an account should be reassigned, or if there is background context Karen should know about, please note it in the REP_NOTE column."),
        ("", "Your notes sync automatically into the Google Sheet and carry forward every month. You only need to fill them in once."),
        ("", ""),
        ("STATUS KEY", ""),
        ("", "🔴 High Value – Lost         Account with $100K+ lifetime revenue, now inactive"),
        ("", "🟠 Declining – At Risk       Account with $20K–$100K lifetime revenue, now inactive"),
        ("", "⚫ Long Dormant              Account inactive for 2+ years"),
        ("", "🟡 Monitor                   Smaller account, inactive — keep an eye on it"),
        ("", "🔴 Stalled                   Active account with zero purchases so far in 2026"),
        ("", "🔴 Severe Decline            Active account on pace to drop 50%+ vs last year"),
        ("", "🟠 Moderate Decline          Active account on pace to drop 20–50% vs last year"),
    ]
    COLUMN_LABELS = {'REP_NOTE', 'FOLLOW_UP_DATE', 'REP_STATUS', 'REMOVE_FLAG'}
    for i, (label, text) in enumerate(instructions, start=3):
        if label in COLUMN_LABELS:
            c = ws_instr.cell(row=i, column=1, value=label)
            c.font = Font(bold=True, color=C_WHITE, size=10, name="Arial")
            c.fill = PatternFill("solid", fgColor=C_MED_BLUE)
            c.alignment = Alignment(vertical="center")
        elif label:
            c = ws_instr.cell(row=i, column=1, value=label)
            c.font = Font(bold=True, color=C_DARK_BLUE, size=10, name="Arial")
        if text:
            c2 = ws_instr.cell(row=i, column=2, value=text)
            c2.font = Font(size=9, name="Arial", color="000000")
            c2.alignment = Alignment(wrap_text=True)
        ws_instr.row_dimensions[i].height = 16

    ws_instr.column_dimensions['A'].width = 24
    ws_instr.column_dimensions['B'].width = 90

    # ── Sheet 2: Inactive Accounts ────────────────────────────────────────────
    ws_inact = wb.create_sheet("🔴 Inactive Accounts")
    ws_inact.sheet_view.showGridLines = False
    ws_inact.freeze_panes = "A4"

    # Title
    n_year_cols = len(year_list)
    last_col = get_column_letter(5 + n_year_cols + 3)
    ws_inact.merge_cells(f"A1:{last_col}1")
    t = ws_inact["A1"]
    t.value = f"INACTIVE ACCOUNTS  —  {rep}"
    t.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_inact.row_dimensions[1].height = 28

    ws_inact.merge_cells(f"A2:{last_col}2")
    sub = ws_inact["A2"]
    sub.value = (
        f"{len(inactive_df)} accounts with no purchases since "
        f"{(TODAY.replace(year=TODAY.year-1)).strftime('%B %d, %Y')}  |  "
        f"Total historical revenue: ${inactive_df['TOTAL'].sum():,.0f}  |  "
        f"Use the Notes / Action column to record your next step"
    )
    sub.font = Font(bold=False, color=C_WHITE, size=9, name="Arial")
    sub.fill = PatternFill("solid", fgColor=C_MED_BLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_inact.row_dimensions[2].height = 16

    # Headers
    headers = (
        ["Account Name", "Status", "Last Sale Date", "Days Inactive"]
        + [str(y) + " ($)" for y in year_list]
        + ["2026 Projected ($)", "Trend vs 2025", f"Total 2020–{max(year_list)} ($)", "Notes / Action"]
    )
    for col_idx, h in enumerate(headers, start=1):
        hdr_cell(ws_inact, 3, col_idx, h)

    # Data rows
    for row_idx, (_, r) in enumerate(inactive_df.sort_values('TOTAL', ascending=False).iterrows(), start=4):
        bg = C_LIGHT_GRAY if row_idx % 2 == 0 else C_WHITE
        data_cell(ws_inact, row_idx, 1, r['CUSTOMER_NAME'], bold=True, bg=bg)

        status = r.get('STATUS', '')
        sc, fc = status_color(status)
        sc_cell = ws_inact.cell(row=row_idx, column=2, value=status)
        sc_cell.font = Font(bold=True, color=fc or C_WHITE, size=9, name="Arial")
        if sc:
            sc_cell.fill = PatternFill("solid", fgColor=sc)
        sc_cell.alignment = Alignment(horizontal="center", vertical="center")

        last_date = r.get('LAST_SALE_DATE')
        data_cell(ws_inact, row_idx, 3,
                  last_date.strftime('%m/%d/%Y') if pd.notna(last_date) and hasattr(last_date, 'strftime') else str(last_date or ''),
                  bg=bg, align="center")
        data_cell(ws_inact, row_idx, 4, int(r.get('DAYS_INACTIVE', 0) or 0), bg=bg, align="center")

        for col_offset, y in enumerate(year_list):
            val = r.get(f"Y{y}", 0)
            val = float(val) if pd.notna(val) and val != '' else 0.0
            c = data_cell(ws_inact, row_idx, 5 + col_offset, val if val else None,
                          bg=bg, num_fmt='$#,##0.00', align="right")

        col_offset2 = len(year_list)
        data_cell(ws_inact, row_idx, 5 + col_offset2, None, bg=bg, align="center")   # 2026 projected
        data_cell(ws_inact, row_idx, 6 + col_offset2, None, bg=bg, align="center")   # trend
        total_val = float(r.get('TOTAL', 0) or 0)
        data_cell(ws_inact, row_idx, 7 + col_offset2, total_val if total_val else None,
                  bold=True, bg=bg, num_fmt='$#,##0.00', align="right")
        note_val = r.get('REP_NOTE', '')
        data_cell(ws_inact, row_idx, 8 + col_offset2, note_val if note_val else '', bg=C_YELLOW)

    # Column widths
    ws_inact.column_dimensions['A'].width = 44
    ws_inact.column_dimensions['B'].width = 26
    ws_inact.column_dimensions['C'].width = 14
    ws_inact.column_dimensions['D'].width = 14
    for i in range(n_year_cols + 3):
        ws_inact.column_dimensions[get_column_letter(5 + i)].width = 14
    ws_inact.column_dimensions[get_column_letter(8 + n_year_cols)].width = 40  # notes

    # ── Sheet 3: Declining Accounts ───────────────────────────────────────────
    ws_decl = wb.create_sheet("📈 Active – Declining")
    ws_decl.sheet_view.showGridLines = False
    ws_decl.freeze_panes = "A4"

    ws_decl.merge_cells(f"A1:{last_col}1")
    t = ws_decl["A1"]
    t.value = f"ACTIVE BUT DECLINING  —  {rep}"
    t.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
    t.fill = PatternFill("solid", fgColor="7B2C2C")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_decl.row_dimensions[1].height = 28

    ws_decl.merge_cells(f"A2:{last_col}2")
    sub = ws_decl["A2"]
    sub.value = (
        f"{len(declining_df)} active accounts with declining purchase volume  |  "
        f"Act now before these become inactive"
    )
    sub.font = Font(bold=False, color=C_WHITE, size=9, name="Arial")
    sub.fill = PatternFill("solid", fgColor="A94040")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_decl.row_dimensions[2].height = 16

    decl_headers = (
        ["Account Name", "Status", "Last Sale Date", "Days Since Last Order"]
        + [str(y) + " ($)" for y in year_list]
        + ["2026 Projected ($)", "Trend vs 2025", f"Total 2020–{max(year_list)} ($)", "Notes / Action"]
    )
    for col_idx, h in enumerate(decl_headers, start=1):
        hdr_cell(ws_decl, 3, col_idx, h, bg="7B2C2C")

    for row_idx, (_, r) in enumerate(declining_df.sort_values('TOTAL', ascending=False).iterrows(), start=4):
        bg = C_LIGHT_GRAY if row_idx % 2 == 0 else C_WHITE
        data_cell(ws_decl, row_idx, 1, r['CUSTOMER_NAME'], bold=True, bg=bg)

        status = r.get('DECLINE_STATUS', '')
        sc_cell = ws_decl.cell(row=row_idx, column=2, value=status)
        sc_cell.font = Font(bold=True, color=C_WHITE, size=9, name="Arial")
        sc_cell.fill = PatternFill("solid", fgColor="C00000")
        sc_cell.alignment = Alignment(horizontal="center", vertical="center")

        last_date = r.get('LAST_SALE_DATE')
        data_cell(ws_decl, row_idx, 3,
                  last_date.strftime('%m/%d/%Y') if pd.notna(last_date) and hasattr(last_date, 'strftime') else str(last_date or ''),
                  bg=bg, align="center")
        data_cell(ws_decl, row_idx, 4, int(r.get('DAYS_INACTIVE', 0) or 0), bg=bg, align="center")

        for col_offset, y in enumerate(year_list):
            val = r.get(f"Y{y}", 0)
            val = float(val) if pd.notna(val) and val != '' else 0.0
            data_cell(ws_decl, row_idx, 5 + col_offset, val if val else None,
                      bg=bg, num_fmt='$#,##0.00', align="right")

        col_offset2 = len(year_list)
        proj_2026 = r.get('PROJ_2026', None)
        trend_pct = r.get('TREND_VS_2025', None)
        data_cell(ws_decl, row_idx, 5 + col_offset2,
                  round(float(proj_2026), 2) if pd.notna(proj_2026) and proj_2026 else None,
                  bg=bg, num_fmt='$#,##0.00', align="right")
        data_cell(ws_decl, row_idx, 6 + col_offset2,
                  round(float(trend_pct), 4) if pd.notna(trend_pct) and trend_pct else None,
                  bg=bg, num_fmt='0%', align="center")
        total_val = float(r.get('TOTAL', 0) or 0)
        data_cell(ws_decl, row_idx, 7 + col_offset2, total_val if total_val else None,
                  bold=True, bg=bg, num_fmt='$#,##0.00', align="right")
        note_val = r.get('REP_NOTE', '')
        data_cell(ws_decl, row_idx, 8 + col_offset2, note_val if note_val else '', bg=C_YELLOW)

    ws_decl.column_dimensions['A'].width = 44
    ws_decl.column_dimensions['B'].width = 26
    ws_decl.column_dimensions['C'].width = 14
    ws_decl.column_dimensions['D'].width = 14
    for i in range(n_year_cols + 3):
        ws_decl.column_dimensions[get_column_letter(5 + i)].width = 14
    ws_decl.column_dimensions[get_column_letter(8 + n_year_cols)].width = 40

    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
#  DATA PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def classify_status(total, days_inactive, last_sale_year):
    """Assign a status flag to an inactive account."""
    if days_inactive is None:
        return '🟡 Monitor'
    days = int(days_inactive)
    if days > 730:
        return '⚫ Long Dormant'
    if total >= HIGH_VALUE_THRESH:
        return '🔴 High Value – Lost'
    if total >= AT_RISK_THRESH:
        return '🟠 Declining – At Risk'
    return '🟡 Monitor'


def classify_decline(proj_2026, rev_2025, rev_2024):
    """Assign a decline status to an active but declining account."""
    if proj_2026 is None or (rev_2025 == 0 and rev_2024 == 0):
        return None, None
    baseline = rev_2025 if rev_2025 > 0 else rev_2024
    if baseline == 0:
        return None, None
    pct_change = (proj_2026 - baseline) / baseline
    if proj_2026 == 0 or pct_change <= -0.99:
        return '🔴 Stalled', pct_change
    if pct_change <= -0.50:
        return '🔴 Severe Decline', pct_change
    if pct_change <= -0.20:
        return '🟠 Moderate Decline', pct_change
    return None, None


def build_summary(df, notes):
    """Aggregate invoice rows into one row per account with year columns and status flags."""
    year_list = sorted(df['INVOICE_YEAR'].dropna().unique().astype(int).tolist())

    # Pivot: revenue per customer per year
    pivot = df.pivot_table(
        index=['CUSTOMER_NAME', 'SALES_REP'],
        columns='INVOICE_YEAR',
        values='INVOICE_TOTAL',
        aggfunc='sum',
        fill_value=0
    ).reset_index()
    pivot.columns = [str(c) if isinstance(c, int) else c for c in pivot.columns]
    # Rename year columns to Y-prefix to avoid Excel treating them as dates
    for y in year_list:
        if str(y) in pivot.columns:
            pivot.rename(columns={str(y): f"Y{y}"}, inplace=True)

    # Last sale info
    last_sale = df.groupby(['CUSTOMER_NAME', 'SALES_REP']).agg(
        LAST_SALE_DATE=('INVOICE_DATE', 'max'),
        TOTAL=('INVOICE_TOTAL', 'sum')
    ).reset_index()
    last_sale['LAST_SALE_DATE'] = pd.to_datetime(last_sale['LAST_SALE_DATE']).dt.date
    last_sale['DAYS_INACTIVE']  = last_sale['LAST_SALE_DATE'].apply(
        lambda d: (TODAY - d).days if pd.notna(d) else None
    )

    summary = last_sale.merge(pivot, on=['CUSTOMER_NAME', 'SALES_REP'], how='left')

    # Status flags
    summary['IS_INACTIVE'] = summary['DAYS_INACTIVE'] >= INACTIVE_DAYS
    summary['STATUS'] = summary.apply(
        lambda r: classify_status(
            r['TOTAL'],
            r['DAYS_INACTIVE'],
            r['LAST_SALE_DATE'].year if r['LAST_SALE_DATE'] else None
        ) if r['IS_INACTIVE'] else 'Active',
        axis=1
    )

    # Declining logic (active accounts only)
    cur_year = TODAY.year
    y_cur  = f"Y{cur_year}"
    y_prev = f"Y{cur_year - 1}"
    y_prev2 = f"Y{cur_year - 2}"
    days_so_far = (TODAY - date(cur_year, 1, 1)).days
    annualized_factor = 365 / max(days_so_far, 1)

    def get_val(r, col):
        v = r.get(col, 0)
        return float(v) if pd.notna(v) and v != '' else 0.0

    summary['PROJ_2026'] = summary.apply(
        lambda r: get_val(r, y_cur) * annualized_factor if not r['IS_INACTIVE'] else None,
        axis=1
    )
    summary['TREND_VS_2025'] = None
    summary['DECLINE_STATUS'] = None
    summary['IS_DECLINING'] = False

    for idx, r in summary.iterrows():
        if r['IS_INACTIVE']:
            continue
        proj   = r.get('PROJ_2026')
        rev25  = get_val(r, y_prev)
        rev24  = get_val(r, y_prev2)
        dstatus, pct = classify_decline(proj, rev25, rev24)
        if dstatus:
            summary.at[idx, 'DECLINE_STATUS'] = dstatus
            summary.at[idx, 'TREND_VS_2025']  = pct
            summary.at[idx, 'IS_DECLINING']   = True

    # Merge rep notes + preserved admin columns (HIDE_FROM_POOL)
    for nc in PRESERVE_COLS:
        summary[nc] = ''
    for idx, r in summary.iterrows():
        key = (r['CUSTOMER_NAME'], r['SALES_REP'])
        if key in notes:
            for nc in PRESERVE_COLS:
                summary.at[idx, nc] = notes[key].get(nc, '')

    return summary, year_list


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════


def push_unowned_accounts_tab(service, h_tabs, collections_df, sid):
    """Push accounts with sales history but no rep attribution to _UNOWNED_ACCOUNTS tab."""
    try:
        import pyodbc as _pyodbc
        conn3 = _pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=localhost;DATABASE=CCCRM;Trusted_Connection=yes;"
        )
        unowned_sql = """
            SELECT TOP 2000
                c.ID               AS CUSTOMER_ID,
                RTRIM(c.NAME)      AS CUSTOMER_NAME,
                RTRIM(c.USERNAME)  AS CRM_USERNAME,
                RTRIM(c.CITY)      AS CITY,
                RTRIM(c.STATE)     AS STATE,
                ISNULL(SUM(i.INVOICE_TOTAL), 0) AS LIFETIME_REVENUE,
                MAX(i.INVOICE_DATE) AS LAST_SALE_DATE
            FROM dbo.CUSTOMER c
            LEFT JOIN dbo.INVCHEA i ON i.CUSTOMER_ID = c.ID
                AND i.INVOICE_DATE >= '2010-01-01'
            WHERE c.INACTIVE = 0
              AND c.ID NOT IN (
                  SELECT arv.ACCOUNT
                  FROM dbo.ACCOUNT_REP_VIEW arv
                  JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                  WHERE e.LOGIN_ID IN (
                      'CKaren','BillP','PIan','RMauricio',
                      'LMancera','bcastor','FJohn','Anolan'
                  )
              )
            GROUP BY c.ID, c.NAME, c.USERNAME, c.CITY, c.STATE
            HAVING ISNULL(SUM(i.INVOICE_TOTAL), 0) > 0
            ORDER BY LIFETIME_REVENUE DESC
        """
        import pandas as _pd2
        df = _pd2.read_sql(unowned_sql, conn3)
        conn3.close()

        ensure_tab_exists(service, '_UNOWNED_ACCOUNTS', h_tabs, sheet_id=sid)
        headers = ['CUSTOMER_ID','CUSTOMER_NAME','CRM_USERNAME','CITY','STATE',
                   'LIFETIME_REVENUE','LAST_SALE_DATE']
        rows = [headers]
        for _, r in df.iterrows():
            def cl(v):
                if v is None: return ''
                s = str(v)
                return '' if s in ('nan','NaT','None') else s
            rows.append([cl(r.get('CUSTOMER_ID')), cl(r.get('CUSTOMER_NAME')),
                         cl(r.get('CRM_USERNAME')), cl(r.get('CITY')), cl(r.get('STATE')),
                         round(float(r.get('LIFETIME_REVENUE') or 0), 2),
                         cl(r.get('LAST_SALE_DATE'))[:10] if r.get('LAST_SALE_DATE') else ''])
        service.spreadsheets().values().clear(
            spreadsheetId=sid, range="'_UNOWNED_ACCOUNTS'!A1:H5000"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sid, range="'_UNOWNED_ACCOUNTS'!A1",
            valueInputOption="RAW", body={"values": rows}
        ).execute()
        _format_header_tab(service, '_UNOWNED_ACCOUNTS', sid=sid)
        print(f"  → Pushed {len(rows)-1} unowned accounts to _UNOWNED_ACCOUNTS tab (History sheet).")
    except Exception as e:
        print(f"  WARNING: Could not push _UNOWNED_ACCOUNTS tab — {e}")


def main():
    print("=" * 60)
    print(f"  Intransit Sales Report  |  {MONTH_STR}")
    print("=" * 60)

    # ── 1. Connect to SQL ─────────────────────────────────────────────────────
    print("\n[1/5] Connecting to CCCRM...")
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=localhost;"
        "DATABASE=CCCRM;"
        "Trusted_Connection=yes;"
    )
    sql = """
        WITH owned_accounts AS (
            SELECT arv.ACCOUNT AS CUSTOMER_ID,
                   RTRIM(e.LOGIN_ID) AS SALES_REP,
                   ROW_NUMBER() OVER (PARTITION BY arv.ACCOUNT ORDER BY e.LOGIN_ID) AS rn
            FROM dbo.ACCOUNT_REP_VIEW arv
            JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
            WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
        )
        SELECT
            i.INVOICE_DATE,
            i.INVOICE_TOTAL,
            oa.SALES_REP,
            c.ID                AS CUSTOMER_ID,
            c.NAME              AS CUSTOMER_NAME
        FROM dbo.INVCHEA i
        JOIN dbo.ORDERHEA o    ON o.ORDER_NUMBER = i.ORDER_NUMBER
        JOIN dbo.CUSTOMER c    ON c.ID = o.CUSTOMER_ID
        JOIN owned_accounts oa ON oa.CUSTOMER_ID = c.ID AND oa.rn = 1
        WHERE i.INVOICE_DATE >= '2010-01-01'
    """
    df = pd.read_sql(sql, conn)

    # Pull primary contact per customer from CONTACT table
    contact_sql = """
        SELECT
            ct.CUSTOMER_ID,
            ct.FIRST_NAME,
            ct.LAST_NAME,
            ct.TITLE,
            ct.EMAIL,
            ct.PHONE,
            ct.MOBIL_PHONE
        FROM dbo.CONTACT ct
        WHERE ct.DEAD = 'N'
          AND ct.CUSTOMER_ID IN (
              SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
              JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
              WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
          )
    """
    contacts_df = pd.read_sql(contact_sql, conn)

    # Build primary contact per customer — prefer contacts with email, then take first
    contacts_df['FULL_NAME']  = (
        contacts_df['FIRST_NAME'].fillna('').str.strip() + ' ' +
        contacts_df['LAST_NAME'].fillna('').str.strip()
    ).str.strip()
    contacts_df['EMAIL']      = contacts_df['EMAIL'].fillna('').str.strip()
    contacts_df['PHONE']      = contacts_df['PHONE'].fillna('').str.strip()
    contacts_df['MOBIL_PHONE']= contacts_df['MOBIL_PHONE'].fillna('').str.strip()
    contacts_df['TITLE']      = contacts_df['TITLE'].fillna('').str.strip()
    contacts_df['CUSTOMER_ID']= pd.to_numeric(contacts_df['CUSTOMER_ID'], errors='coerce')

    # Sort so contacts with email come first, then pick one per customer
    contacts_df['HAS_EMAIL'] = contacts_df['EMAIL'].str.len() > 0
    contacts_df = contacts_df.sort_values(['CUSTOMER_ID', 'HAS_EMAIL'], ascending=[True, False])
    primary_contacts = contacts_df.drop_duplicates(subset='CUSTOMER_ID', keep='first')[[
        'CUSTOMER_ID', 'FULL_NAME', 'TITLE', 'EMAIL', 'PHONE', 'MOBIL_PHONE'
    ]].rename(columns={
        'FULL_NAME':   'CONTACT_NAME',
        'TITLE':       'CONTACT_TITLE',
        'EMAIL':       'CONTACT_EMAIL',
        'PHONE':       'CONTACT_PHONE',
        'MOBIL_PHONE': 'CONTACT_MOBILE',
    })
    print(f"  → Pulled {len(primary_contacts):,} primary contacts.")

    df['INVOICE_DATE']  = pd.to_datetime(df['INVOICE_DATE'])
    df['INVOICE_YEAR']  = df['INVOICE_DATE'].dt.year
    df['INVOICE_TOTAL'] = pd.to_numeric(df['INVOICE_TOTAL'], errors='coerce').fillna(0)
    df['SALES_REP']     = df['SALES_REP'].astype(str).str.strip()
    print(f"  → Pulled {len(df):,} invoice rows.")
    print(f"  → Reps found: {sorted(df['SALES_REP'].unique().tolist())}")

    # ── Pull ALL contacts (not just primary) ─────────────────────────────────
    print("  → Pulling all contacts...")
    all_contacts_sql = """
        SELECT
            ct.CUSTOMER_ID,
            c.NAME              AS CUSTOMER_NAME,
            (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
             JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
             WHERE arv2.ACCOUNT = c.ID
               AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
             ORDER BY e2.LOGIN_ID) AS SALES_REP,
            ct.FIRST_NAME,
            ct.LAST_NAME,
            ct.TITLE,
            ct.EMAIL,
            ct.PHONE,
            ct.MOBIL_PHONE
        FROM dbo.CONTACT ct
        JOIN dbo.CUSTOMER c ON c.ID = ct.CUSTOMER_ID
        WHERE ct.DEAD = 'N'
          AND ct.CUSTOMER_ID IN (
              SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
              JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
              WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
          )
        ORDER BY ct.CUSTOMER_ID, ct.LAST_NAME
    """
    all_contacts_df = pd.read_sql(all_contacts_sql, conn)
    all_contacts_df['CONTACT_NAME'] = (
        all_contacts_df['FIRST_NAME'].fillna('').str.strip() + ' ' +
        all_contacts_df['LAST_NAME'].fillna('').str.strip()
    ).str.strip()
    all_contacts_df['TITLE']       = all_contacts_df['TITLE'].fillna('').str.strip()
    all_contacts_df['EMAIL']       = all_contacts_df['EMAIL'].fillna('').str.strip()
    all_contacts_df['PHONE']       = all_contacts_df['PHONE'].fillna('').str.strip()
    all_contacts_df['MOBILE']      = all_contacts_df['MOBIL_PHONE'].fillna('').str.strip()
    all_contacts_df['CUSTOMER_ID'] = pd.to_numeric(all_contacts_df['CUSTOMER_ID'], errors='coerce')
    print(f"  → Pulled {len(all_contacts_df):,} total contacts.")

    # ── Pull invoice line items (part numbers) for last 2 years ──────────────
    print("  → Pulling invoice line items...")
    line_items_sql = """
        WITH owned_accounts AS (
            SELECT arv.ACCOUNT AS CUSTOMER_ID,
                   RTRIM(e.LOGIN_ID) AS SALES_REP,
                   ROW_NUMBER() OVER (PARTITION BY arv.ACCOUNT ORDER BY e.LOGIN_ID) AS rn
            FROM dbo.ACCOUNT_REP_VIEW arv
            JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
            WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
        )
        SELECT TOP 100000
            i.INVOICE_NUMBER,
            i.INVOICE_DATE,
            i.INVOICE_TOTAL,
            oa.SALES_REP,
            c.ID                AS CUSTOMER_ID,
            c.NAME              AS CUSTOMER_NAME,
            d.FULLPART          AS PART_NUMBER,
            d.DESCRIPTION,
            d.QTY_INVOICED      AS QTY,
            d.PRICE             AS UNIT_PRICE,
            d.EXTENDED_PRICE    AS LINE_TOTAL,
            d.CONDITION
        FROM dbo.INVCHEA i
        JOIN dbo.ORDERHEA o    ON o.ORDER_NUMBER = i.ORDER_NUMBER
        JOIN dbo.CUSTOMER c    ON c.ID = o.CUSTOMER_ID
        JOIN dbo.INVCDTL d     ON d.INVOICE_NUMBER = i.INVOICE_NUMBER
        JOIN owned_accounts oa ON oa.CUSTOMER_ID = c.ID AND oa.rn = 1
        WHERE i.INVOICE_DATE >= '2010-01-01'
          AND d.FULLPART IS NOT NULL
          AND d.FULLPART != ''
        ORDER BY i.INVOICE_DATE DESC
    """
    try:
        line_items_df = pd.read_sql(line_items_sql, conn)
        line_items_df['INVOICE_DATE'] = pd.to_datetime(line_items_df['INVOICE_DATE']).dt.date.astype(str)
        line_items_df['INVOICE_TOTAL'] = pd.to_numeric(line_items_df['INVOICE_TOTAL'], errors='coerce').fillna(0)
        line_items_df['UNIT_PRICE']    = pd.to_numeric(line_items_df['UNIT_PRICE'],    errors='coerce').fillna(0)
        line_items_df['LINE_TOTAL']    = pd.to_numeric(line_items_df['LINE_TOTAL'],    errors='coerce').fillna(0)
        line_items_df['QTY']           = pd.to_numeric(line_items_df['QTY'],           errors='coerce').fillna(0)
        print(f"  → Pulled {len(line_items_df):,} invoice line items.")
    except Exception as e:
        print(f"  WARNING: Could not pull line items — {e}")
        line_items_df = pd.DataFrame()

    # ── Pull collections data — balances, payments, notes ────────────────────
    print("  → Pulling collections data...")
    try:
        collections_sql = """
            WITH owned_accounts AS (
                SELECT arv.ACCOUNT AS CUSTOMER_ID,
                       RTRIM(e.LOGIN_ID) AS SALES_REP,
                       ROW_NUMBER() OVER (PARTITION BY arv.ACCOUNT ORDER BY e.LOGIN_ID) AS rn
                FROM dbo.ACCOUNT_REP_VIEW arv
                JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
            )
            SELECT
                i.CUSTOMER_ID,
                c.NAME              AS CUSTOMER_NAME,
                oa.SALES_REP,
                SUM(i.BALANCE)      AS CURRENT_BALANCE,
                MAX(cb.CREDIT_LIMIT) AS CREDIT_LIMIT,
                c.AVERAGE_PAY,
                c.CREDIT_HOLD,
                c.TERMS             AS PAYMENT_TERMS,
                MAX(i.INVOICE_DATE) AS LAST_INVOICE_DATE,
                NULL                AS LAST_PAYMENT_DATE,
                NULL                AS LAST_PAYMENT_AMOUNT,
                NULL                AS COLLECTION_NOTE,
                NULL                AS NOTE_DATE,
                NULL                AS NOTE_BY
            FROM dbo.INVCHEA i
            JOIN dbo.CUSTOMER c    ON c.ID = i.CUSTOMER_ID
            JOIN owned_accounts oa ON oa.CUSTOMER_ID = c.ID AND oa.rn = 1
            LEFT JOIN dbo.CUSTBAL cb ON cb.CUSTOMER_ID = i.CUSTOMER_ID
            WHERE i.BALANCE > 0
              AND i.INVOICE_DATE >= '2020-01-01'
            GROUP BY i.CUSTOMER_ID, c.NAME, oa.SALES_REP,
                     c.AVERAGE_PAY, c.CREDIT_HOLD, c.TERMS
        """
        collections_df = pd.read_sql(collections_sql, conn)

        # Try to pull last payment per customer
        try:
            pay_sql = """
                SELECT
                    p.CUSTOMER_ID,
                    MAX(p.DATE)              AS LAST_PAYMENT_DATE,
                    SUM(p.AMT)               AS LAST_PAYMENT_AMOUNT,
                    AVG(CAST(p.DAYS_OVER AS FLOAT)) AS AVG_DAYS_OVER
                FROM dbo.PAY p
                WHERE p.DATE >= DATEADD(year, -2, GETDATE())
                GROUP BY p.CUSTOMER_ID
            """
            pay_df = pd.read_sql(pay_sql, conn)
            pay_df['CUSTOMER_ID'] = pd.to_numeric(pay_df['CUSTOMER_ID'], errors='coerce')
            collections_df['CUSTOMER_ID'] = pd.to_numeric(collections_df['CUSTOMER_ID'], errors='coerce')
            collections_df = collections_df.merge(
                pay_df[['CUSTOMER_ID','LAST_PAYMENT_DATE','LAST_PAYMENT_AMOUNT']],
                on='CUSTOMER_ID', how='left', suffixes=('','_PAY')
            )
            if 'LAST_PAYMENT_DATE_PAY' in collections_df.columns:
                collections_df['LAST_PAYMENT_DATE'] = collections_df['LAST_PAYMENT_DATE_PAY']
                collections_df['LAST_PAYMENT_AMOUNT'] = collections_df['LAST_PAYMENT_AMOUNT_PAY']
            if 'AVG_DAYS_OVER_PAY' in collections_df.columns:
                collections_df['AVG_DAYS_OVER'] = collections_df['AVG_DAYS_OVER_PAY']
            elif 'AVG_DAYS_OVER' not in collections_df.columns:
                collections_df['AVG_DAYS_OVER'] = 0
        except Exception as e2:
            print(f"    (Payment history unavailable: {e2})")

        # Try to pull collection notes
        try:
            collnote_sql = """
                SELECT
                    cn.CUSTOMER_ID,
                    cn.NOTE             AS COLLECTION_NOTE,
                    cn.DATE_TIME        AS NOTE_DATE,
                    cn.USERNAME         AS NOTE_BY,
                    cn.SUBJECT
                FROM dbo.COLLNOTE cn
                ORDER BY cn.DATE_TIME DESC
            """
            collnote_df = pd.read_sql(collnote_sql, conn)
            collnote_df['CUSTOMER_ID'] = pd.to_numeric(collnote_df['CUSTOMER_ID'], errors='coerce')
            # Keep most recent note per customer
            collnote_df = collnote_df.drop_duplicates(subset='CUSTOMER_ID', keep='first')
            collections_df['CUSTOMER_ID'] = pd.to_numeric(collections_df['CUSTOMER_ID'], errors='coerce')
            collections_df = collections_df.merge(
                collnote_df[['CUSTOMER_ID','COLLECTION_NOTE','NOTE_DATE','NOTE_BY']],
                on='CUSTOMER_ID', how='left', suffixes=('','_CN')
            )
            for col in ['COLLECTION_NOTE','NOTE_DATE','NOTE_BY']:
                if col + '_CN' in collections_df.columns:
                    collections_df[col] = collections_df[col + '_CN']
        except Exception as e3:
            print(f"    (Collection notes unavailable: {e3})")

        print(f"  → Pulled {len(collections_df):,} accounts with open balances.")
    except Exception as e:
        print(f"  WARNING: Could not pull collections data — {e}")
        collections_df = pd.DataFrame()

    # ── Pull RFQ (Reqs) history ───────────────────────────────────────────────
    print("  → Pulling RFQ history...")
    try:
        reqs_sql = """
        SELECT TOP 200000
            r.RECNUM            AS REQ_ID,
            r.CONTACT_ID,
            ct.CUSTOMER_ID,
            c.NAME              AS CUSTOMER_NAME,
            RTRIM(r.SELLER)     AS SALES_REP,
            r.DATE_ENTERED,
            r.FULLPART          AS PART_NUMBER,
            r.MFR               AS MANUFACTURER,
            r.QTY,
            r.TARGET_PRICE,
            r.DESCRIPTION,
            r.CONDITION,
            r.CLOSED,
            r.DATE_CLOSED,
            r.SELLER
        FROM dbo.REQ r
        LEFT JOIN dbo.CONTACT ct ON ct.RECNUM = r.CONTACT_ID
        LEFT JOIN dbo.CUSTOMER c  ON c.ID = ct.CUSTOMER_ID
        WHERE r.DATE_ENTERED >= DATEADD(year, -1, GETDATE())
          AND RTRIM(r.SELLER) IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
        ORDER BY r.DATE_ENTERED DESC
    """
        reqs_df = pd.read_sql(reqs_sql, conn)
        reqs_df['DATE_ENTERED'] = pd.to_datetime(reqs_df['DATE_ENTERED']).dt.date.astype(str)
        reqs_df['DATE_CLOSED']  = pd.to_datetime(reqs_df['DATE_CLOSED'], errors='coerce').dt.date.astype(str)
        reqs_df['TARGET_PRICE'] = pd.to_numeric(reqs_df['TARGET_PRICE'], errors='coerce').fillna(0)
        reqs_df['QTY']          = pd.to_numeric(reqs_df['QTY'], errors='coerce').fillna(0)
        print(f"  → Pulled {len(reqs_df):,} RFQ records.")
    except Exception as e:
        print(f"  WARNING: Could not pull RFQ data — {e}")
        reqs_df = pd.DataFrame()

    # ── Pull Quote history ────────────────────────────────────────────────────
    print("  → Pulling quote history...")
    try:
        quotes_sql = """
            SELECT TOP 100000
                q.ID                AS QUOTE_ID,
                q.CONTACT_ID,
                c.ID                AS CUSTOMER_ID,
                c.NAME              AS CUSTOMER_NAME,
                (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
                 JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
                 WHERE arv2.ACCOUNT = c.ID
                   AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
                 ORDER BY e2.LOGIN_ID) AS SALES_REP,
                q.QUOTE_DATE,
                q.CLOSED,
                q.Username          AS ENTERED_BY,
                ql.FULLPART         AS PART_NUMBER,
                ql.MFR              AS MANUFACTURER,
                ql.QTY,
                ql.PRICE,
                ql.DESCRIPTION,
                ql.CONDITION,
                ql.CLOSED           AS LINE_CLOSED
            FROM dbo.QUOTE_HD q
            JOIN dbo.CONTACT ct  ON ct.RECNUM = q.CONTACT_ID
            JOIN dbo.CUSTOMER c  ON c.ID = ct.CUSTOMER_ID
            JOIN dbo.QUOTE_LI ql ON ql.ID = q.ID
            WHERE q.QUOTE_DATE >= DATEADD(year, -1, GETDATE())
              AND c.ID IN (
                  SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
                  JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                  WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
              )
            ORDER BY q.QUOTE_DATE DESC
        """
        quotes_df = pd.read_sql(quotes_sql, conn)
        quotes_df['QUOTE_DATE'] = pd.to_datetime(quotes_df['QUOTE_DATE'], errors='coerce').dt.date.astype(str)
        quotes_df['PRICE']      = pd.to_numeric(quotes_df['PRICE'], errors='coerce').fillna(0)
        quotes_df['QTY']        = pd.to_numeric(quotes_df['QTY'], errors='coerce').fillna(0)
        print(f"  → Pulled {len(quotes_df):,} quote line records.")
    except Exception as e:
        print(f"  WARNING: Could not pull quote data — {e}")
        quotes_df = pd.DataFrame()

    # ── Pull customer stats from CUSTOMER table ───────────────────────────────
    print("  → Pulling customer stats...")
    try:
        stats_sql = """
            SELECT
                c.ID                AS CUSTOMER_ID,
                c.NAME              AS CUSTOMER_NAME,
                (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
                 JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
                 WHERE arv2.ACCOUNT = c.ID
                   AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
                 ORDER BY e2.LOGIN_ID) AS SALES_REP,
                c.ESTABLISHED,
                c.NUM_OF_BUYERS,
                c.NUM_OF_EMP,
                c.REVENUE,
                c.OPEN_REQS,
                c.OPEN_QUOTES,
                c.OPEN_ORDERS,
                c.Reqs              AS TOTAL_REQS,
                c.Quotes            AS TOTAL_QUOTES,
                c.Invoices          AS TOTAL_INVOICES,
                c.ORDERS            AS TOTAL_ORDERS,
                c.RETURNS           AS TOTAL_RETURNS,
                c.LY                AS LY_SALES,
                c.AVERAGE_PAY,
                c.BALANCE_S         AS CURRENT_BALANCE,
                c.CREDIT_LINE_S     AS CREDIT_LIMIT,
                c.CREDIT_HOLD,
                c.LAST_REQ,
                c.LAST_QUOTE,
                c.LAST_ORDER,
                c.LAST_INVOICE,
                c.SEGMENT,
                c.AcctClass         AS ACCOUNT_CLASS,
                c.TYPE              AS ACCOUNT_TYPE,
                c.LAST_ACTIVITY,
                c.LAST_CONTACTED
            FROM dbo.CUSTOMER c
            WHERE c.ID IN (
                SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
                JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
            )
        """
        cust_stats_df = pd.read_sql(stats_sql, conn)
        cust_stats_df['CUSTOMER_ID'] = pd.to_numeric(cust_stats_df['CUSTOMER_ID'], errors='coerce')
        print(f"  → Pulled stats for {len(cust_stats_df):,} customers.")
    except Exception as e:
        print(f"  WARNING: Could not pull customer stats — {e}")
        cust_stats_df = pd.DataFrame()

    # ── Pull open sales orders (booked not yet shipped) ─────────────────────
    print("  → Pulling open orders (booked)...")
    try:
        orders_sql = """
            SELECT
                o.ORDER_NUMBER,
                o.CUSTOMER_ID,
                c.NAME              AS CUSTOMER_NAME,
                (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
                 JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
                 WHERE arv2.ACCOUNT = o.CUSTOMER_ID
                   AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
                 ORDER BY e2.LOGIN_ID) AS SALES_REP,
                o.ORDER_DATE,
                d.DETAIL_NUMBER,
                d.FULLPART          AS PART_NUMBER,
                d.DESCRIPTION,
                d.QTY_ORDERED,
                d.QTY_SHIPPED,
                d.QTY_NEEDED,
                d.PRICE,
                d.EXTENDED_PRICE,
                d.COST,
                d.CONDITION,
                d.CLOSED
            FROM dbo.ORDERHEA o
            JOIN dbo.CUSTOMER c   ON c.ID = o.CUSTOMER_ID
            JOIN dbo.ORDERDTL d   ON d.ORDER_NUMBER = o.ORDER_NUMBER
            WHERE o.CUSTOMER_ID IN (
                SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
                JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
            )
              AND d.CLOSED = 'N'
              AND d.QTY_NEEDED > 0
            ORDER BY o.ORDER_DATE DESC
        """
        orders_df = pd.read_sql(orders_sql, conn)
        orders_df['ORDER_DATE']     = pd.to_datetime(orders_df['ORDER_DATE'], errors='coerce').dt.date.astype(str)
        orders_df['EXTENDED_PRICE'] = pd.to_numeric(orders_df['EXTENDED_PRICE'], errors='coerce').fillna(0)
        orders_df['PRICE']          = pd.to_numeric(orders_df['PRICE'],          errors='coerce').fillna(0)
        orders_df['COST']           = pd.to_numeric(orders_df['COST'],           errors='coerce').fillna(0)
        orders_df['QTY_ORDERED']    = pd.to_numeric(orders_df['QTY_ORDERED'],    errors='coerce').fillna(0)
        orders_df['QTY_SHIPPED']    = pd.to_numeric(orders_df['QTY_SHIPPED'],    errors='coerce').fillna(0)
        orders_df['QTY_NEEDED']     = pd.to_numeric(orders_df['QTY_NEEDED'],     errors='coerce').fillna(0)
        print(f"  → Pulled {len(orders_df):,} open order lines (booked).")
    except Exception as e:
        print(f"  WARNING: Could not pull open orders — {e}")
        orders_df = pd.DataFrame()

    # ── Pull GP from INVCDTL ──────────────────────────────────────────────────
    print("  → Pulling GP data from invoices...")
    try:
        gp_sql = """
            WITH owned_accounts AS (
                SELECT arv.ACCOUNT AS CUSTOMER_ID,
                       RTRIM(e.LOGIN_ID) AS SALES_REP,
                       ROW_NUMBER() OVER (PARTITION BY arv.ACCOUNT ORDER BY e.LOGIN_ID) AS rn
                FROM dbo.ACCOUNT_REP_VIEW arv
                JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
            )
            SELECT
                i.INVOICE_DATE,
                oa.SALES_REP,
                c.NAME              AS CUSTOMER_NAME,
                i.INVOICE_NUMBER,
                i.INVOICE_TOTAL     AS EXTENDED_PRICE,
                i.GP,
                NULL                AS PART_NUMBER,
                MONTH(i.INVOICE_DATE) AS INVOICE_MONTH,
                YEAR(i.INVOICE_DATE)  AS INVOICE_YEAR,
                o.ORDERED_BY        AS BUYER_NAME
            FROM dbo.INVCHEA i
            JOIN dbo.ORDERHEA o    ON o.ORDER_NUMBER = i.ORDER_NUMBER
            JOIN dbo.CUSTOMER c    ON c.ID = o.CUSTOMER_ID
            JOIN owned_accounts oa ON oa.CUSTOMER_ID = c.ID AND oa.rn = 1
            WHERE i.INVOICE_DATE >= '2010-01-01'
        """
        gp_df = pd.read_sql(gp_sql, conn)
        gp_df['INVOICE_DATE']   = pd.to_datetime(gp_df['INVOICE_DATE'], errors='coerce').dt.date.astype(str)
        gp_df['EXTENDED_PRICE'] = pd.to_numeric(gp_df['EXTENDED_PRICE'], errors='coerce').fillna(0)
        gp_df['GP']             = pd.to_numeric(gp_df['GP'],             errors='coerce').fillna(0)
        print(f"  → Pulled {len(gp_df):,} GP records.")
    except Exception as e:
        print(f"  WARNING: Could not pull GP data — {e}")
        gp_df = pd.DataFrame()


    # Pull orders history (all orders with dates, for booked/dashboard calc)
    print("  -> Pulling orders history...")
    try:
        ord_hist_sql = """
            SELECT
                o.ORDER_NUMBER,
                RTRIM(c.NAME)               AS CUSTOMER_NAME,
                (SELECT TOP 1 RTRIM(e2.LOGIN_ID) FROM dbo.ACCOUNT_REP_VIEW arv2
                 JOIN dbo.EMPLOYEE e2 ON e2.ID = arv2.SALES_REP
                 WHERE arv2.ACCOUNT = o.CUSTOMER_ID
                   AND e2.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
                 ORDER BY e2.LOGIN_ID) AS SALES_REP,
                CONVERT(VARCHAR(10), o.ORDER_DATE, 120) AS ORDER_DATE,
                d.FULLPART          AS PART_NUMBER,
                d.DESCRIPTION,
                d.EXTENDED_PRICE,
                d.COST,
                d.CONDITION
            FROM dbo.ORDERHEA o
            JOIN dbo.CUSTOMER c   ON c.ID = o.CUSTOMER_ID
            JOIN dbo.ORDERDTL d   ON d.ORDER_NUMBER = o.ORDER_NUMBER
            WHERE o.ORDER_DATE >= DATEADD(year, -2, GETDATE())
              AND o.CUSTOMER_ID IN (
                  SELECT arv.ACCOUNT FROM dbo.ACCOUNT_REP_VIEW arv
                  JOIN dbo.EMPLOYEE e ON e.ID = arv.SALES_REP
                  WHERE e.LOGIN_ID IN ('CKaren','BillP','PIan','RMauricio','LMancera','bcastor','FJohn','Anolan')
              )
        """
        ord_hist_df = pd.read_sql(ord_hist_sql, conn)
        ord_hist_df['EXTENDED_PRICE'] = pd.to_numeric(ord_hist_df['EXTENDED_PRICE'], errors='coerce').fillna(0)
        ord_hist_df['COST']           = pd.to_numeric(ord_hist_df['COST'],           errors='coerce').fillna(0)
        ord_hist_df['GP']             = ord_hist_df['EXTENDED_PRICE'] - ord_hist_df['COST']
        print(f"  -> Pulled {len(ord_hist_df):,} order history records.")
    except Exception as e:
        print(f"  WARNING: Could not pull orders history -- {e}")
        ord_hist_df = pd.DataFrame()

    conn.close()
    print("  → SQL connection closed.")

    # ── 2. Connect to Sheets + backup notes BEFORE touching anything ─────────────
    print("\n[2/7] Connecting to Google Sheet & backing up rep notes...")
    service = get_sheets_service()
    if service:
        meta     = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        all_tabs = [s['properties']['title'] for s in meta['sheets']]
        backup_notes_to_sheet(service, all_tabs)
    else:
        all_tabs = []

    # ── 3. Read notes back to carry forward ──────────────────────────────────────
    print("\n[3/7] Reading rep notes from Google Sheet...")
    notes = read_notes_from_sheet(service)

    # ── 4. Build account summary ────────────────────────────────────────────────
    print("\n[4/7] Building account summary...")
    summary, year_list = build_summary(df, notes)

    # Merge primary contact info into summary
    # Need CUSTOMER_ID on summary — pull it from df
    cust_id_map = df[['CUSTOMER_NAME','CUSTOMER_ID']].copy()
    cust_id_map['CUSTOMER_NAME'] = cust_id_map['CUSTOMER_NAME'].astype(str).str.strip()
    cust_id_map['CUSTOMER_ID']   = pd.to_numeric(cust_id_map['CUSTOMER_ID'], errors='coerce')
    cust_id_map = cust_id_map.drop_duplicates(subset='CUSTOMER_NAME')
    summary['CUSTOMER_NAME_CLEAN'] = summary['CUSTOMER_NAME'].astype(str).str.strip()
    summary = summary.merge(cust_id_map, left_on='CUSTOMER_NAME_CLEAN', right_on='CUSTOMER_NAME', how='left', suffixes=('','_MAP'))
    summary = summary.merge(primary_contacts, on='CUSTOMER_ID', how='left')
    for col in ['CONTACT_NAME','CONTACT_TITLE','CONTACT_EMAIL','CONTACT_PHONE','CONTACT_MOBILE']:
        summary[col] = summary[col].fillna('')
    summary = summary.drop(columns=['CUSTOMER_NAME_CLEAN', 'CUSTOMER_NAME_MAP'], errors='ignore')
    print(f"  → {len(summary)} total accounts built.")
    matched = (summary['CONTACT_EMAIL'].str.len() > 0).sum()
    print(f"  → {matched} accounts have a contact email address.")

    # Remove accounts approved by John only
    removed = summary[summary['JOHN_APPROVAL'].str.strip().str.upper() == 'YES']
    if len(removed):
        print(f"  → Removing {len(removed)} accounts approved for removal by John")
    summary = summary[summary['JOHN_APPROVAL'].str.strip().str.upper() != 'YES']

    # ── 5. Generate Excel files ────────────────────────────────────────────────
    print("\n[5/7] Generating Excel reports...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for rep in ACTIVE_REPS:
        rep_df = summary[summary['SALES_REP'] == rep]
        inactive_count  = rep_df['IS_INACTIVE'].sum()
        declining_count = rep_df['IS_DECLINING'].sum()
        if len(rep_df) == 0:
            print(f"  Rep {rep}: no accounts — skipping")
            continue
        print(f"  Rep {rep}: {inactive_count} inactive, {declining_count} declining")
        out_path = os.path.join(OUTPUT_DIR, f"AccountReview_{rep}_{MONTH_STR}.xlsx")
        build_rep_excel(rep, rep_df, out_path, year_list)

    # ── 6. Push to Google Sheet ───────────────────────────────────────────────────
    print("\n[6/7] Pushing data to Google Sheet...")
    if service:
        meta     = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        all_tabs = [s['properties']['title'] for s in meta['sheets']]

        for rep in ACTIVE_REPS:
            rep_df = summary[summary['SALES_REP'] == rep]
            if len(rep_df) > 0:
                push_rep_tab(service, rep, rep_df, year_list, all_tabs)

        # push_master_tab skipped — MASTER tab not used by app, freeing Main sheet cells

        # Push new data tabs
        if len(all_contacts_df) > 0:
            push_contacts_tab(service, all_contacts_df, all_tabs)
        if len(line_items_df) > 0:
            push_invoice_history_tab(service, line_items_df, all_tabs)
        if len(collections_df) > 0:
            push_collections_tab(service, collections_df, all_tabs)
        push_collections_invoices_tab(service, all_tabs)
        if len(reqs_df) > 0:
            push_reqs_tab(service, reqs_df, all_tabs)
        if len(quotes_df) > 0:
            push_quotes_tab(service, quotes_df, all_tabs)
        if len(cust_stats_df) > 0:
            push_customer_stats_tab(service, cust_stats_df, all_tabs)
        push_customer_directory_tab(service, conn, all_tabs)
        if len(orders_df) > 0:
            push_open_orders_tab(service, orders_df, all_tabs)
        if len(ord_hist_df) > 0:
            push_orders_history_tab(service, ord_hist_df, all_tabs)
        if len(gp_df) > 0:
            print("  → Pausing 20s before GP tab to avoid rate limit...")
            time.sleep(20)
            push_gp_tab(service, gp_df, all_tabs)

        stats = {
            'total':    len(summary),
            'inactive': int(summary['IS_INACTIVE'].sum()),
            'at_risk':  int((summary['STATUS'].str.contains('At Risk', na=False)).sum()),
            'active':   int((~summary['IS_INACTIVE']).sum()),
            'note':     f'{len(removed)} accounts removed by John'
        }
        log_run_to_sheet(service, stats, all_tabs)
    else:
        print("  → Google Sheets skipped (no connection).")

    inactive_total  = int(summary['IS_INACTIVE'].sum())
    declining_total = int(summary['IS_DECLINING'].sum())
    print("\n" + "=" * 60)
    print("  Done!  All reports generated.")
    print(f"  Total accounts   : {len(summary)}")
    print(f"  Inactive         : {inactive_total}")
    print(f"  Declining        : {declining_total}")
    print(f"  Active           : {len(summary) - inactive_total}")
    print(f"  Total contacts   : {len(all_contacts_df)}")
    print(f"  Invoice lines    : {len(line_items_df)}")
    print(f"  Open balances    : {len(collections_df)}")
    print(f"  RFQ records      : {len(reqs_df)}")
    print(f"  Quote records    : {len(quotes_df)}")
    print(f"  Customer stats   : {len(cust_stats_df)}")
    print(f"  Open orders      : {len(orders_df)}")
    print(f"  Orders history   : {len(ord_hist_df)}")
    print(f"  GP records       : {len(gp_df)}")
    print(f"  Excel files in   : {OUTPUT_DIR}")
    print(f"  Google Sheet     : https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}")
    print("=" * 60)


if __name__ == "__main__":
    main()